# -*- mode:python; coding:utf-8 -*-
# Copyright (c) 2024 IBM Corp. All rights reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     https://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
"""Transform POAM spreadsheet to OSCAL POAM JSON format."""

# mypy: ignore-errors  # noqa E800
import configparser
import datetime
import logging
import pathlib
import re
import traceback
import uuid
from typing import Any, Dict, Iterator, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from trestle.oscal import OSCAL_VERSION
from trestle.oscal.common import (
    Metadata,
    Observation,
    Origin,
    OriginActor,
    Property,
    RelatedObservation,
    RelatedRisk,
    Response,
    Risk,
    RiskStatus,
    SubjectReference,
    SystemId,
    Task as OscalTask,
    Timing,
    WithinDateRange,
)
from trestle.oscal.poam import PlanOfActionAndMilestones, PoamItem
from trestle.tasks.base_task import TaskBase, TaskOutcome

logger = logging.getLogger(__name__)


class UUIDManager:
    """Manage deterministic UUID generation for POAM objects."""

    # Namespace UUID for this task
    NAMESPACE = uuid.UUID('e8d8efc6-c23e-4e3e-a2e8-bc8fc08ff6c3')

    @staticmethod
    def poam_item_uuid(poam_id: str) -> str:
        """Generate deterministic UUID for PoamItem from POAM ID."""
        return str(uuid.uuid5(UUIDManager.NAMESPACE, f'poam-item-{poam_id}'))

    @staticmethod
    def observation_uuid(poam_id: str) -> str:
        """Generate deterministic UUID for Observation from POAM ID."""
        return str(uuid.uuid5(UUIDManager.NAMESPACE, f'observation-{poam_id}'))

    @staticmethod
    def risk_uuid(poam_id: str) -> str:
        """Generate deterministic UUID for Risk from POAM ID."""
        return str(uuid.uuid5(UUIDManager.NAMESPACE, f'risk-{poam_id}'))

    @staticmethod
    def task_uuid(poam_id: str, milestone_index: int) -> str:
        """Generate deterministic UUID for Task (milestone) from POAM ID + index."""
        return str(uuid.uuid5(UUIDManager.NAMESPACE, f'task-{poam_id}-{milestone_index}'))

    @staticmethod
    def actor_uuid(actor_name: str) -> str:
        """Generate deterministic UUID for Origin Actor."""
        return str(uuid.uuid5(UUIDManager.NAMESPACE, f'actor-{actor_name}'))


class PoamValidator:
    """Validate POAM spreadsheet data before transformation."""

    VALID_RISK_RATINGS = ['Low', 'Moderate', 'High', 'N/A', '']
    VALID_YES_NO_PENDING = ['Yes', 'No', 'Pending', '']
    CONTROL_PATTERN = re.compile(r'^[A-Z]{2}-\d+(\(\d+\))?$')

    def __init__(self, validate_mode: str = 'warn') -> None:
        """
        Initialize validator.

        Args:
            validate_mode: 'on' (fail on error), 'warn' (log warnings), or 'off' (skip validation)
        """
        self.validate_mode = validate_mode
        self.errors = []
        self.warnings = []

    def validate_row(self, row_num: int, row_data: Dict[str, Any]) -> List[str]:
        """
        Validate a single row of data.

        Args:
            row_num: Row number in spreadsheet
            row_data: Dictionary of column_name -> value

        Returns:
            List of validation error messages
        """
        errors = []

        # Required fields
        if not row_data.get('POAM ID'):
            errors.append(f'Row {row_num}: Missing required field "POAM ID"')
        if not row_data.get('Weakness Name'):
            errors.append(f'Row {row_num}: Missing required field "Weakness Name"')
        if not row_data.get('Weakness Description'):
            errors.append(f'Row {row_num}: Missing required field "Weakness Description"')
        if not row_data.get('Controls'):
            errors.append(f'Row {row_num}: Missing required field "Controls"')

        # Risk rating validation
        for rating_field in ['Original Risk Rating', 'Adjusted Risk Rating']:
            value = row_data.get(rating_field, '')
            if value and value not in self.VALID_RISK_RATINGS:
                errors.append(
                    f'Row {row_num}: Invalid {rating_field}: "{value}". '
                    f'Must be one of: {", ".join(self.VALID_RISK_RATINGS[:-1])}'
                )

        # Yes/No/Pending field validation
        for field in ['Risk Adjustment', 'False Positive', 'Operational Requirement']:
            value = row_data.get(field, '')
            if value and value not in self.VALID_YES_NO_PENDING:
                errors.append(
                    f'Row {row_num}: Invalid {field}: "{value}". '
                    f'Must be one of: {", ".join(self.VALID_YES_NO_PENDING[:-1])}'
                )

        return errors

    def parse_controls(self, controls_str: str) -> List[str]:
        """
        Parse and validate control IDs.

        Args:
            controls_str: Comma/space-separated control IDs like "AC-1, AC-2, SC-7(5)"

        Returns:
            List of validated control IDs
        """
        if not controls_str:
            return []

        # Split by comma and/or space
        controls = re.split(r'[,\s]+', controls_str.strip())

        validated = []
        for ctrl in controls:
            ctrl = ctrl.strip()
            if not ctrl:
                continue
            if self.CONTROL_PATTERN.match(ctrl.upper()):
                validated.append(ctrl.upper())
            else:
                if self.validate_mode != 'off':
                    logger.warning(f'Invalid control format: "{ctrl}" (expected format: XX-N or XX-N(N))')

        return validated

    def log_validation_results(self) -> bool:
        """
        Log validation results based on validation mode.

        Returns:
            True if validation passed or mode is 'warn'/'off', False if errors in 'on' mode
        """
        if self.validate_mode == 'off':
            return True

        if self.errors:
            if self.validate_mode == 'on':
                for error in self.errors:
                    logger.error(error)
                return False
            else:  # warn mode
                for error in self.errors:
                    logger.warning(error)

        if self.warnings:
            for warning in self.warnings:
                logger.warning(warning)

        return True


class PoamXlsxHelper:
    """Helper class for reading POAM spreadsheet templates."""

    # Column name constants
    POAM_ID = 'POAM ID'
    CONTROLS = 'Controls'
    WEAKNESS_NAME = 'Weakness Name'
    WEAKNESS_DESCRIPTION = 'Weakness Description'
    WEAKNESS_DETECTOR_SOURCE = 'Weakness Detector Source'
    WEAKNESS_SOURCE_IDENTIFIER = 'Weakness Source Identifier'
    ASSET_IDENTIFIER = 'Asset Identifier'
    POINT_OF_CONTACT = 'Point of Contact'
    RESOURCES_REQUIRED = 'Resources Required'
    OVERALL_REMEDIATION_PLAN = 'Overall Remediation Plan'
    ORIGINAL_DETECTION_DATE = 'Original Detection Date'
    SCHEDULED_COMPLETION_DATE = 'Scheduled Completion Date'
    PLANNED_MILESTONES = 'Planned Milestones'
    MILESTONE_CHANGES = 'Milestone Changes'
    STATUS_DATE = 'Status Date'
    VENDOR_DEPENDENCY = 'Vendor Dependency'
    LAST_VENDOR_CHECKIN_DATE = 'Last Vendor Check-in Date'
    VENDOR_DEPENDENT_PRODUCT_NAME = 'Vendor Dependent Product Name'
    ORIGINAL_RISK_RATING = 'Original Risk Rating'
    ADJUSTED_RISK_RATING = 'Adjusted Risk Rating'
    RISK_ADJUSTMENT = 'Risk Adjustment'
    FALSE_POSITIVE = 'False Positive'
    OPERATIONAL_REQUIREMENT = 'Operational Requirement'
    DEVIATION_RATIONALE = 'Deviation Rationale'
    SUPPORTING_DOCUMENTS = 'Supporting Documents'
    COMMENTS = 'Comments'
    AUTO_APPROVE = 'Auto-Approve'
    BOD_22_01_TRACKING = 'Binding Operational Directive 22-01 tracking'
    BOD_22_01_DUE_DATE = 'Binding Operational Directive 22-01 Due Date'
    CVE = 'CVE'
    SERVICE_NAME = 'Service Name'

    def __init__(self) -> None:
        """Initialize helper."""
        self._column_map: Dict[str, int] = {}
        self._work_sheet = None
        self._header_row = 5  # Template has headers at row 5 (1-indexed)

    def load(self, xlsx_path: pathlib.Path, sheet_name: str = 'Open POA&M Items') -> None:
        """
        Load spreadsheet file and map columns.

        Args:
            xlsx_path: Path to spreadsheet file
            sheet_name: Name of worksheet to load

        Raises:
            FileNotFoundError: If file doesn't exist
            KeyError: If worksheet doesn't exist
        """
        if not xlsx_path.exists():
            raise FileNotFoundError(f'Spreadsheet file not found: {xlsx_path}')

        workbook = load_workbook(filename=str(xlsx_path), data_only=True)

        if sheet_name not in workbook.sheetnames:
            available = ', '.join(workbook.sheetnames)
            raise KeyError(f'Worksheet "{sheet_name}" not found. Available sheets: {available}')

        self._work_sheet = workbook[sheet_name]
        self._map_columns()

    def _map_columns(self) -> None:
        """Map column names to column indices from header row."""
        if self._work_sheet is None:
            return

        # Read header row (row 5 in template, 1-indexed)
        for cell in self._work_sheet[self._header_row]:
            if cell.value and isinstance(cell.value, str):
                col_name = cell.value.strip()
                if col_name:
                    self._column_map[col_name] = cell.column

        logger.debug(f'Mapped {len(self._column_map)} columns')

    def row_generator(self) -> Iterator[Tuple[int, Dict[str, Any]]]:
        """
        Generate row numbers and data dictionaries for data rows.

        Yields:
            Tuple of (row_number, row_data_dict)
        """
        if self._work_sheet is None:
            return

        # Data starts at row 6 (after header at row 5)
        data_start_row = self._header_row + 1
        max_row = self._work_sheet.max_row

        for row_num in range(data_start_row, max_row + 1):
            row_data = self._get_row_data(row_num)

            # Skip empty rows (no POAM ID)
            if not row_data.get(self.POAM_ID):
                continue

            yield row_num, row_data

    def _get_row_data(self, row_num: int) -> Dict[str, Any]:
        """
        Extract data from a row as dictionary.

        Args:
            row_num: Row number (1-indexed)

        Returns:
            Dictionary mapping column names to cell values
        """
        row_data = {}

        for col_name, col_idx in self._column_map.items():
            cell = self._work_sheet.cell(row=row_num, column=col_idx)
            value = None if isinstance(cell, MergedCell) else cell.value
            row_data[col_name] = self._clean_value(value)

        return row_data

    def _clean_value(self, value: Any) -> Any:
        """
        Clean cell value.

        Args:
            value: Raw cell value

        Returns:
            Cleaned value (strings are stripped, None for empty)
        """
        if value is None:
            return None
        if isinstance(value, str):
            value = value.strip()
            return value if value else None
        return value

    def parse_date(self, date_value: Any) -> Optional[datetime.datetime]:
        """
        Parse spreadsheet date to datetime with timezone.

        Args:
            date_value: Spreadsheet date (datetime object or string)

        Returns:
            datetime with UTC timezone or None
        """
        if date_value is None:
            return None

        if isinstance(date_value, datetime.datetime):
            # Add timezone if missing
            if date_value.tzinfo is None:
                return date_value.replace(tzinfo=datetime.timezone.utc)
            return date_value

        if isinstance(date_value, datetime.date):
            # Convert date to datetime
            dt = datetime.datetime.combine(date_value, datetime.time.min)
            return dt.replace(tzinfo=datetime.timezone.utc)

        if isinstance(date_value, str):
            # Try to parse ISO 8601 format
            try:
                dt = datetime.datetime.fromisoformat(date_value.replace('Z', '+00:00'))
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=datetime.timezone.utc)
                return dt
            except (ValueError, AttributeError):
                logger.warning(f'Could not parse date string: "{date_value}"')
                return None

        logger.warning(f'Unexpected date type: {type(date_value)}')
        return None

    def parse_milestones(self, milestones_str: str) -> List[Dict[str, Any]]:
        """
        Parse milestone text into structured format.

        Args:
            milestones_str: Milestone text (may contain multiple milestones)

        Returns:
            List of milestone dictionaries with 'title', 'description', optional 'timing'
        """
        if not milestones_str:
            return []

        milestones = []
        lines = milestones_str.split('\n')

        for line in lines:
            line = line.strip()
            if not line:
                continue

            # Try to parse: "Milestone N: Description [by YYYY-MM-DD]"
            match = re.match(
                r'(Milestone\s+\d+|M\d+)[:.]?\s*(.+?)(?:\s+by\s+(\d{4}-\d{2}-\d{2}))?$', line, re.IGNORECASE
            )
            if match:
                milestone_num, description, date_str = match.groups()
                milestone = {'title': description.strip(), 'description': milestone_num.strip()}
                if date_str:
                    milestone['timing'] = date_str
                milestones.append(milestone)
            else:
                # Fallback: treat entire line as milestone title
                milestones.append({'title': line, 'description': 'Milestone'})

        return milestones


class PoamBuilder:
    """Builder class for constructing OSCAL POAM objects from spreadsheet data."""

    def __init__(self, timestamp: str, validator: PoamValidator) -> None:
        """
        Initialize builder.

        Args:
            timestamp: ISO timestamp for metadata
            validator: PoamValidator instance
        """
        self._timestamp = timestamp
        self._validator = validator

    def create_poam_item(self, poam_id: str, row_data: Dict[str, Any]) -> PoamItem:
        """
        Create PoamItem from row data.

        Args:
            poam_id: POAM ID
            row_data: Row data dictionary

        Returns:
            PoamItem object
        """
        # Required fields
        weakness_name = row_data.get(PoamXlsxHelper.WEAKNESS_NAME, '')
        title = weakness_name.strip() if weakness_name else ''
        weakness_desc = row_data.get(PoamXlsxHelper.WEAKNESS_DESCRIPTION, '')
        description = weakness_desc.strip() if weakness_desc else ''

        # Optional fields
        comments = row_data.get(PoamXlsxHelper.COMMENTS)
        if comments and isinstance(comments, str):
            comments = comments.strip() if comments.strip() else None

        # Build properties
        props = []

        # Add POAM ID as property (clean it first)
        clean_poam_id = poam_id.strip() if poam_id else poam_id
        if clean_poam_id:
            props.append(Property(name='poam-id', value=clean_poam_id))

        # Add control IDs as properties
        controls_str = row_data.get(PoamXlsxHelper.CONTROLS, '')
        if controls_str:
            controls = self._validator.parse_controls(controls_str)
            for ctrl_id in controls:
                props.append(Property(name='control-id', value=ctrl_id))

        # Create PoamItem
        poam_item = PoamItem(
            uuid=UUIDManager.poam_item_uuid(poam_id), title=title, description=description, props=props or None
        )

        # Add remarks if present
        if comments:
            poam_item.remarks = comments

        return poam_item

    def create_observation(self, poam_id: str, row_data: Dict[str, Any], helper: PoamXlsxHelper) -> Observation:
        """
        Create Observation from row data.

        Args:
            poam_id: POAM ID
            row_data: Row data dictionary
            helper: PoamXlsxHelper instance

        Returns:
            Observation object
        """
        weakness_name_raw = row_data.get(PoamXlsxHelper.WEAKNESS_NAME, '')
        weakness_name = weakness_name_raw.strip() if weakness_name_raw else ''

        # Description is required
        description = f'Weakness detected: {weakness_name}'
        weakness_source_id = row_data.get(PoamXlsxHelper.WEAKNESS_SOURCE_IDENTIFIER)
        if weakness_source_id and isinstance(weakness_source_id, str):
            weakness_source_id = weakness_source_id.strip()
            if weakness_source_id:
                description += f' (Source: {weakness_source_id})'

        # Collected date (required)
        detection_date_value = row_data.get(PoamXlsxHelper.ORIGINAL_DETECTION_DATE)
        collected = helper.parse_date(detection_date_value)
        if collected is None:
            # Default to current timestamp if no date provided
            collected = datetime.datetime.fromisoformat(self._timestamp)

        # Methods are required
        methods = ['TEST']

        # Origins are optional
        origins = None
        detector_source = row_data.get(PoamXlsxHelper.WEAKNESS_DETECTOR_SOURCE)
        if detector_source:
            actor = OriginActor(type='tool', actor_uuid=UUIDManager.actor_uuid(detector_source))
            origin = Origin(actors=[actor])
            origins = [origin]

        # Subjects are optional
        subjects = None
        asset_id = row_data.get(PoamXlsxHelper.ASSET_IDENTIFIER)
        if asset_id:
            subject = SubjectReference(subject_uuid=UUIDManager.actor_uuid(asset_id), type='component')
            subjects = [subject]

        observation = Observation(
            uuid=UUIDManager.observation_uuid(poam_id),
            description=description,
            methods=methods,
            collected=collected,
            origins=origins,
            subjects=subjects,
        )

        return observation

    def create_risk(self, poam_id: str, row_data: Dict[str, Any], helper: PoamXlsxHelper) -> Risk:
        """
        Create Risk from row data.

        Args:
            poam_id: POAM ID
            row_data: Row data dictionary
            helper: PoamXlsxHelper instance

        Returns:
            Risk object
        """
        # Required fields - clean all text fields
        weakness_name = row_data.get(PoamXlsxHelper.WEAKNESS_NAME, '')
        title = weakness_name.strip() if weakness_name else ''
        weakness_desc = row_data.get(PoamXlsxHelper.WEAKNESS_DESCRIPTION, '')
        description = weakness_desc.strip() if weakness_desc else ''
        statement_raw = row_data.get(PoamXlsxHelper.OVERALL_REMEDIATION_PLAN, description)
        if statement_raw and isinstance(statement_raw, str):
            statement = statement_raw.strip()
        else:
            statement = statement_raw if statement_raw else description
        status = RiskStatus(__root__='open')  # Default status for Open POA&M Items sheet

        # Properties
        props = []

        # Helper function to clean and validate property values
        def add_property_if_valid(name: str, value: Any) -> None:
            """Add property if value is valid (non-empty after stripping)."""
            if value is None:
                return

            if isinstance(value, str):
                # Strip all whitespace including newlines, tabs, etc.
                cleaned = value.strip()
                # Additional check: ensure not just whitespace and matches OSCAL pattern
                if cleaned and not cleaned.isspace():
                    try:
                        props.append(Property(name=name, value=cleaned))
                    except Exception as e:
                        logger.warning(f'Could not create property {name} with value "{cleaned[:50]}...": {e}')
            elif value:
                # Non-string value, convert to string
                props.append(Property(name=name, value=str(value)))

        # Risk ratings as properties
        add_property_if_valid('original-risk-rating', row_data.get(PoamXlsxHelper.ORIGINAL_RISK_RATING))
        add_property_if_valid('adjusted-risk-rating', row_data.get(PoamXlsxHelper.ADJUSTED_RISK_RATING))
        add_property_if_valid('risk-adjustment', row_data.get(PoamXlsxHelper.RISK_ADJUSTMENT))
        add_property_if_valid('false-positive', row_data.get(PoamXlsxHelper.FALSE_POSITIVE))
        add_property_if_valid('operational-requirement', row_data.get(PoamXlsxHelper.OPERATIONAL_REQUIREMENT))
        add_property_if_valid('deviation-rationale', row_data.get(PoamXlsxHelper.DEVIATION_RATIONALE))

        # Deadline is optional
        deadline = None
        completion_date_value = row_data.get(PoamXlsxHelper.SCHEDULED_COMPLETION_DATE)
        if completion_date_value:
            deadline = helper.parse_date(completion_date_value)

        # Remediations with milestones (optional)
        remediations = None
        milestones_str = row_data.get(PoamXlsxHelper.PLANNED_MILESTONES)
        if milestones_str:
            milestones = helper.parse_milestones(milestones_str)
            if milestones:
                tasks = self._create_milestone_tasks(poam_id, milestones, helper)
                remediation = Response(
                    uuid=str(uuid.uuid4()),
                    lifecycle='planned',
                    title=f'Remediation for {poam_id}',
                    description=statement,
                    tasks=tasks or None,
                )
                remediations = [remediation]

        risk = Risk(
            uuid=UUIDManager.risk_uuid(poam_id),
            title=title,
            description=description,
            statement=statement,
            status=status,
            props=props or None,
            deadline=deadline,
            remediations=remediations,
        )

        return risk

    def _create_milestone_tasks(self, poam_id: str, milestones: List[Dict[str, Any]],
                                helper: PoamXlsxHelper) -> List[OscalTask]:
        """
        Create OSCAL Task objects from milestone data.

        Args:
            poam_id: POAM ID
            milestones: List of milestone dictionaries
            helper: PoamXlsxHelper instance

        Returns:
            List of OscalTask objects
        """
        tasks = []

        for idx, milestone in enumerate(milestones):
            title = milestone.get('title', '')
            description = milestone.get('description')

            # Timing is optional
            timing = None
            date_str = milestone.get('timing')
            if date_str:
                try:
                    end_date = helper.parse_date(date_str)
                    if end_date:
                        # Create a date range (start = now, end = milestone date)
                        start_date = datetime.datetime.fromisoformat(self._timestamp)
                        timing = Timing(within_date_range=WithinDateRange(start=start_date, end=end_date))
                except Exception as e:
                    logger.warning(f'Could not parse milestone date "{date_str}": {e}')

            task = OscalTask(
                uuid=UUIDManager.task_uuid(poam_id, idx),
                type='milestone',
                title=title,
                description=description,
                timing=timing,
            )
            tasks.append(task)

        return tasks

    def link_objects(self, poam_item: PoamItem, observation: Observation, risk: Risk) -> None:
        """
        Link POAM objects together via UUID references.

        Args:
            poam_item: PoamItem to link
            observation: Observation to link
            risk: Risk to link
        """
        # Link PoamItem to Observation
        poam_item.related_observations = [RelatedObservation(observation_uuid=observation.uuid)]

        # Link PoamItem to Risk
        poam_item.related_risks = [RelatedRisk(risk_uuid=risk.uuid)]

        # Link Risk to Observation
        risk.related_observations = [RelatedObservation(observation_uuid=observation.uuid)]


class XlsxToOscalPoam(TaskBase):
    """
    Transform POAM spreadsheet to OSCAL POAM JSON.

    This task reads a POAM spreadsheet template (specifically the
    "Open POA&M Items" worksheet) and transforms each row into an
    OSCAL Plan of Action and Milestones (POAM) JSON file.

    Each spreadsheet row creates three linked OSCAL objects:
    1. PoamItem: The main weakness/issue description
    2. Observation: Details about when/how the weakness was detected
    3. Risk: Risk assessment, remediation plan, and milestones

    Configuration Example:
        [task.xlsx-to-oscal-poam]
        xlsx-file = POAM-Template.xlsx
        output-dir = output/
        title = MySystem POA&M
        version = 1.0

    Spreadsheet Requirements:
        - Must use POAM template structure
        - Headers at row 5
        - Data starts at row 6
        - Required columns: POAM ID, Weakness Name, Weakness Description, Controls

    See Also:
        - docs/tutorials/task.xlsx-to-oscal-poam.md
        - OSCAL POAM specification
    """

    name = 'xlsx-to-oscal-poam'

    def __init__(self, config_object: Optional[configparser.SectionProxy]) -> None:
        """
        Initialize trestle task xlsx-to-oscal-poam.

        Args:
            config_object: Config section associated with the task.
        """
        super().__init__(config_object)
        self._timestamp = datetime.datetime.now(datetime.timezone.utc).replace(microsecond=0).isoformat()

    def set_timestamp(self, timestamp: str) -> None:
        """Set the timestamp."""
        self._timestamp = timestamp

    def print_info(self) -> None:
        """Print the help string."""
        logger.info(f'Help information for {self.name} task.')
        logger.info('')
        logger.info('Purpose: Transform POAM spreadsheet to OSCAL POAM JSON format.')
        logger.info('')
        logger.info(f'Configuration flags sit under [task.{self.name}]:')
        text1 = '  xlsx-file                = '
        text2 = '(required) the path of the POAM spreadsheet file.'
        logger.info(text1 + text2)
        text1 = '  work-sheet-name          = '
        text2 = '(optional) the name of the work sheet in the spreadsheet file (default: "Open POA&M Items").'
        logger.info(text1 + text2)
        text1 = '  output-dir               = '
        text2 = '(required) the path of the output directory for synthesized OSCAL POAM .json file.'
        logger.info(text1 + text2)
        text1 = '  title                    = '
        text2 = '(required) the title for the POAM.'
        logger.info(text1 + text2)
        text1 = '  version                  = '
        text2 = '(required) the version of the POAM.'
        logger.info(text1 + text2)
        text1 = '  system-id                = '
        text2 = '(optional) the system identifier.'
        logger.info(text1 + text2)
        text1 = '  output-overwrite         = '
        text2 = '(optional) true [default] or false; replace existing output when true.'
        logger.info(text1 + text2)
        text1 = '  validate-required-fields = '
        text2 = '(optional) validation mode for required fields: on, warn [default], or off.'
        logger.info(text1 + text2)
        text1 = '  quiet                    = '
        text2 = '(optional) true or false [default]; suppress info messages when true.'
        logger.info(text1 + text2)
        logger.info('')
        logger.info('Expected columns in spreadsheet:')
        text1 = '                             '
        text2 = 'column "POAM ID" contains unique identifier for each POAM item (required).'
        logger.info(text1 + text2)
        text2 = 'column "Weakness Name" contains title/name of the weakness (required).'
        logger.info(text1 + text2)
        text2 = 'column "Weakness Description" contains description of the weakness (required).'
        logger.info(text1 + text2)
        text2 = 'column "Controls" contains related security control IDs (required).'
        logger.info(text1 + text2)
        text2 = 'column "Weakness Detector Source" contains source that detected the weakness (optional).'
        logger.info(text1 + text2)
        text2 = 'column "Weakness Source Identifier" contains identifier from detection source (optional).'
        logger.info(text1 + text2)
        text2 = 'column "Asset Identifier" contains identifier of affected asset (optional).'
        logger.info(text1 + text2)
        text2 = 'column "Point of Contact" contains contact information for responsible party (optional).'
        logger.info(text1 + text2)
        text2 = 'column "Resources Required" contains resources needed for remediation (optional).'
        logger.info(text1 + text2)
        text2 = 'column "Overall Remediation Plan" contains the remediation plan description (optional).'
        logger.info(text1 + text2)
        text2 = 'column "Original Detection Date" contains date weakness was first detected (optional).'
        logger.info(text1 + text2)
        text2 = 'column "Scheduled Completion Date" contains target completion date (optional).'
        logger.info(text1 + text2)
        text2 = 'column "Planned Milestones" contains milestone descriptions and dates (optional).'
        logger.info(text1 + text2)
        text2 = 'column "Original Risk Rating" contains initial risk assessment (optional).'
        logger.info(text1 + text2)
        text2 = 'column "Adjusted Risk Rating" contains adjusted risk assessment (optional).'
        logger.info(text1 + text2)
        text2 = 'column "Risk Adjustment" contains rationale for risk adjustment (optional).'
        logger.info(text1 + text2)
        text2 = 'column "False Positive" contains yes/no indicator for false positive (optional).'
        logger.info(text1 + text2)
        text2 = 'column "Operational Requirement" contains yes/no indicator for operational requirement (optional).'
        logger.info(text1 + text2)
        text2 = 'column "Deviation Rationale" contains rationale for deviation (optional).'
        logger.info(text1 + text2)
        text2 = 'column "Supporting Documents" contains references to supporting documentation (optional).'
        logger.info(text1 + text2)
        text2 = 'column "Comments" contains additional comments or notes (optional).'
        logger.info(text1 + text2)

    def configure(self) -> bool:
        """
        Configure the task.

        Returns:
            True if configuration successful, False otherwise
        """
        if not self._config:
            logger.error('Config section is missing')
            return False

        # Required parameters
        self._xlsx_file = self._config.get('xlsx-file')
        if not self._xlsx_file:
            logger.error('Missing required parameter: xlsx-file')
            return False

        self._output_dir = self._config.get('output-dir')
        if not self._output_dir:
            logger.error('Missing required parameter: output-dir')
            return False

        self._title = self._config.get('title')
        if not self._title:
            logger.error('Missing required parameter: title')
            return False

        self._version = self._config.get('version')
        if not self._version:
            logger.error('Missing required parameter: version')
            return False

        # Optional parameters
        self._work_sheet_name = self._config.get('work-sheet-name', 'Open POA&M Items')
        self._system_id = self._config.get('system-id')
        self._overwrite = self._config.getboolean('output-overwrite', True)
        self._validate_mode = self._config.get('validate-required-fields', 'warn')
        self._quiet = self._config.getboolean('quiet', False)

        return True

    def simulate(self) -> TaskOutcome:
        """Provide a simulated outcome."""
        return TaskOutcome('simulated-success')

    def execute(self) -> TaskOutcome:
        """Provide an executed outcome."""
        try:
            return self._execute()
        except Exception:
            logger.error(traceback.format_exc())
            return TaskOutcome('failure')

    def _execute(self) -> TaskOutcome:
        """Execute path core."""
        # Configure
        if not self.configure():
            return TaskOutcome('failure')

        # Setup output directory
        output_path = pathlib.Path(self._output_dir)
        output_path.mkdir(exist_ok=True, parents=True)

        # Setup Excel helper
        xlsx_path = pathlib.Path(self._xlsx_file)
        helper = PoamXlsxHelper()

        try:
            helper.load(xlsx_path, self._work_sheet_name)
        except FileNotFoundError as e:
            logger.error(str(e))
            return TaskOutcome('failure')
        except KeyError as e:
            logger.error(str(e))
            return TaskOutcome('failure')

        # Setup validator
        validator = PoamValidator(validate_mode=self._validate_mode)

        # Setup builder
        builder = PoamBuilder(self._timestamp, validator)

        # Process rows
        poam_items = []
        observations = []
        risks = []

        for row_num, row_data in helper.row_generator():
            # Validate row
            errors = validator.validate_row(row_num, row_data)
            if errors:
                validator.errors.extend(errors)
                if self._validate_mode == 'on':
                    continue  # Skip invalid rows in strict mode

            # Extract POAM ID
            poam_id = row_data.get(PoamXlsxHelper.POAM_ID, '')

            # Create OSCAL objects
            poam_item = builder.create_poam_item(poam_id, row_data)
            observation = builder.create_observation(poam_id, row_data, helper)
            risk = builder.create_risk(poam_id, row_data, helper)

            # Link objects
            builder.link_objects(poam_item, observation, risk)

            # Add to lists
            poam_items.append(poam_item)
            observations.append(observation)
            risks.append(risk)

        # Check validation results
        if not validator.log_validation_results():
            logger.error('Validation failed')
            return TaskOutcome('failure')

        if not poam_items:
            logger.error('No valid POAM items found in Excel file')
            return TaskOutcome('failure')

        # Create POAM
        poam = self._create_poam(poam_items, observations, risks)

        # Write output
        output_file = output_path / 'plan-of-action-and-milestones.json'
        if not self._overwrite and output_file.exists():
            logger.error(f'Output file already exists: {output_file}')
            return TaskOutcome('failure')

        poam.oscal_write(output_file)

        if not self._quiet:
            logger.info(f'Created POAM with {len(poam_items)} items')
            logger.info(f'Output: {output_file}')

        return TaskOutcome('success')

    def _create_poam(
        self, poam_items: List[PoamItem], observations: List[Observation], risks: List[Risk]
    ) -> PlanOfActionAndMilestones:
        """
        Create OSCAL PlanOfActionAndMilestones object.

        Args:
            poam_items: List of PoamItem objects
            observations: List of Observation objects
            risks: List of Risk objects

        Returns:
            PlanOfActionAndMilestones object
        """
        # Create metadata
        metadata = Metadata(
            title=self._title, last_modified=self._timestamp, oscal_version=OSCAL_VERSION, version=self._version
        )

        # Optional system-id
        system_id = None
        if self._system_id:
            system_id = SystemId(identifier_type='https://ietf.org/rfc/rfc4122', id=self._system_id)

        # Create POAM
        poam = PlanOfActionAndMilestones(
            uuid=str(uuid.uuid4()),
            metadata=metadata,
            system_id=system_id,
            poam_items=poam_items,
            observations=observations or None,
            risks=risks or None,
        )

        return poam
