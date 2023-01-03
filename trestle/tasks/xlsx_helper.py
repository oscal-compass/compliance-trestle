# -*- mode:python; coding:utf-8 -*-
# Copyright (c) 2022 IBM Corp. All rights reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     https://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
"""XLSX utilities."""

import logging
import pathlib
import string
from typing import Any, Dict, Iterator, List, Tuple

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

from trestle import __version__
from trestle.common.err import TrestleError
from trestle.common.list_utils import is_ordered_sublist
from trestle.core.catalog.catalog_interface import CatalogInterface
from trestle.oscal.catalog import Catalog
from trestle.tasks.base_task import TaskBase

logger = logging.getLogger(__name__)


def get_trestle_version() -> str:
    """Get trestle version wrapper."""
    return __version__


class Column():
    """Spread sheet columns."""

    control_id = 'ControlId'
    control_text = 'ControlText'
    goal_name_id = 'goal_name_id'
    goal_version = 'goal_version'
    rule_name_id = 'rule_name_id'
    rule_version = 'rule_version'
    nist_mappings = 'NIST Mappings'
    resource_title = 'ResourceTitle'
    parameter_opt_parm = 'Parameter [optional parameter]'
    values_alternatives = 'Values default , [alternatives]'
    filter_column = None

    tokens_nist_mappings = nist_mappings.split()
    tokens_parameter_opt_parm = parameter_opt_parm.split()
    rename_parameter_opt_parm = 'ParameterName'
    tokens_values_alternatives = values_alternatives.split()
    rename_values_alternatives = 'ParameterValue'

    help_list = []
    text1 = '                      '
    text2 = f'column "{control_id}" contains control ID.'
    help_list.append(text1 + text2)
    text2 = f'column "{control_text}" contains control text.'
    help_list.append(text1 + text2)
    text2 = f'columns "{nist_mappings}" contain NIST control mappings.'
    help_list.append(text1 + text2)
    text2 = f'column "{resource_title}" contains component name.'
    help_list.append(text1 + text2)
    text2 = f'column "{goal_name_id}" contains goal name.'
    help_list.append(text1 + text2)
    text2 = f'column "{goal_version}" contains goal version.'
    help_list.append(text1 + text2)
    text2 = f'column "{rule_name_id}" contains rule name.'
    help_list.append(text1 + text2)
    text2 = f'column "{rule_version}" contains rule version.'
    help_list.append(text1 + text2)
    text2 = f'column "{parameter_opt_parm}" contains parameter name + description, separated by newline.'
    help_list.append(text1 + text2)
    text2 = f'column "{values_alternatives}" contains parameter values.'
    help_list.append(text1 + text2)


class XlsxHelper:
    """Xlsx Helper common functions and assistance navigating spread sheet."""

    by_goal = 'by-goal'
    by_rule = 'by-rule'
    by_control = 'by-control'
    by_check = 'by-check'

    profile_types = [by_goal, by_rule, by_control, by_check]

    def __init__(self) -> None:
        """Initialize."""
        self._column = Column()

    def print_info(self, name, oscal_name) -> None:
        """Print the help string."""
        logger.info(f'Help information for {name} task.')
        logger.info('')
        logger.info(f'Purpose: From spread sheet and catalog produce OSCAL {oscal_name} file.')
        logger.info('')
        logger.info(f'Configuration flags sit under [task.{name}]:')
        if oscal_name == 'component_definition':
            text1 = '  catalog-file      = '
            text2 = '(required) the path of the OSCAL catalog file.'
            logger.info(text1 + text2)
        text1 = '  spread-sheet-file = '
        text2 = '(required) the path of the spread sheet file.'
        logger.info(text1 + text2)
        text1 = '  work-sheet-name   = '
        text2 = '(required) the name of the work sheet in the spread sheet file.'
        logger.info(text1 + text2)
        for line in self._column.help_list:
            logger.info(line)
        text1 = '  output-dir        = '
        text2 = '(required) the path of the output directory for synthesized OSCAL .json files.'
        logger.info(text1 + text2)
        text1 = '  output-overwrite  = '
        text2 = '(optional) true [default] or false; replace existing output when true.'
        logger.info(text1 + text2)
        text1 = '  filter-column     = '
        text2 = '(optional) column heading of yes/no values; process only "yes" rows.'
        logger.info(text1 + text2)
        text1 = '  profile-type      = '
        text2 = f'(optional) one of {self.profile_types}'
        logger.info(text1 + text2)

    @property
    def profile_type(self) -> str:
        """Profile type."""
        return self._profile_type

    def configure(self, task: TaskBase) -> bool:
        """Configure."""
        if not task._config:
            logger.warning('config missing')
            return False
        # config verbosity
        quiet = task._config.get('quiet', False)
        task._verbose = not quiet
        # required for component-definition
        if not self.configure_cd(task):
            return False
        # required for profile
        if not self.configure_profile(task):
            return False
        # optional
        self._column.filter_column = task._config.get('filter-column', None)
        # config spread sheet
        spread_sheet = task._config.get('spread-sheet-file')
        if spread_sheet is None:
            logger.warning('config missing "spread-sheet"')
            return False
        if not pathlib.Path(spread_sheet).exists():
            logger.warning('"spread-sheet" not found')
            return False
        sheet_name = task._config.get('work-sheet-name')
        if sheet_name is None:
            logger.warning('config missing "work-sheet-name"')
            return False
        # announce spreadsheet
        if task._verbose:
            logger.info(f'input: {spread_sheet}')
        # get profile type
        if task.name == 'xlsx-to-oscal-profile':
            self._profile_type = task._config.get('profile-type', self.profile_types[0])
            if self._profile_type not in self.profile_types:
                logger.warning(f'invalid "profile-type" {self._profile_type} ')
                return False
        else:
            self._profile_type = None
        # load spread sheet
        self.load(spread_sheet, sheet_name)
        return True

    def configure_cd(self, task: TaskBase) -> bool:
        """Configure cd."""
        if task.name == 'xlsx-to-oscal-cd':
            catalog_file = task._config.get('catalog-file')
            if catalog_file is None:
                logger.warning('config missing "catalog-file"')
                return False
            try:
                catalog = Catalog.oscal_read(pathlib.Path(catalog_file))
                logger.debug(f'catalog: {catalog_file}')
            except Exception as e:  # pragma: no cover
                raise TrestleError(f'Error loading catalog {catalog_file}: {e}')
            task.catalog_interface = CatalogInterface(catalog)
        return True

    def configure_profile(self, task: TaskBase) -> bool:
        """Configure profile."""
        if task.name == 'xlsx-to-oscal-profile':
            profile_title = task._config.get('profile-title')
            if profile_title is None:
                logger.warning('config missing "profile-title"')
                return False
            spread_sheet_url = task._config.get('spread-sheet-url')
            if spread_sheet_url is None:
                logger.warning('config missing "spread-sheet-url"')
                return False
        return True

    def load(self, spread_sheet: str, sheet_name: str) -> None:
        """Load."""
        self._spread_sheet = spread_sheet
        self._sheet_name = sheet_name
        self._wb = load_workbook(self._spread_sheet)
        self._work_sheet = self._wb[self._sheet_name]
        self._map_name_to_letters = {}
        # accumulators
        self.rows_missing_control_id = []
        self.rows_missing_goal_name_id = []
        self.rows_invalid_goal_name_id = []
        self.rows_missing_rule_name_id = []
        self.rows_invalid_rule_name_id = []
        self.rows_invalid_parameter_name = []
        self.rows_missing_controls = []
        self.rows_missing_parameters = []
        self.rows_missing_parameters_values = []
        self.rows_filtered = []
        # map columns
        self._map_columns()

    def row_generator(self) -> Iterator[int]:
        """Generate rows until control_id is None."""
        row = 1
        rows_skipped_consecutive = 0
        # assume no more data when 100 consecutve rows no control id
        rows_skipped_consecutive_limit = 100
        while True:
            row = row + 1
            control_id = self._get_control_id(row)
            goal_id = self.get_goal_name_id(row)
            if control_id is None and goal_id is None:
                rows_skipped_consecutive += 1
                if rows_skipped_consecutive < rows_skipped_consecutive_limit:
                    continue
                logger.debug(f'break: {row} {rows_skipped_consecutive}')
                break
            if control_id is None:
                self._add_row(row, self.rows_missing_control_id)
                continue
            if goal_id is None:
                self._add_row(row, self.rows_missing_goal_name_id)
                continue
            if self._is_filtered(row):
                continue
            yield row
            rows_skipped_consecutive = 0

    def _is_filtered(self, row) -> bool:
        """Return True if row is to be skipped."""
        if self._column.filter_column is None:
            return False
        col = self._get_column_letter(self._column.filter_column)
        value = self._work_sheet[col + str(row)].value
        if value is None:
            return False
        if value.lower() != 'yes':
            return False
        self._add_row(row, self.rows_filtered)
        return True

    def get_goal_name_id(self, row: int, strict: bool = True) -> str:
        """Get goal_name_id from work_sheet."""
        col = self._get_column_letter(self._column.goal_name_id)
        value = self._work_sheet[col + str(row)].value
        if value is None:
            self._add_row(row, self.rows_missing_goal_name_id)
        else:
            value = str(value).strip()
            if strict:
                svalue = str(value).strip()
                value = ''.join(str(svalue).split())
                if value != svalue:
                    self._add_row(row, self.rows_invalid_goal_name_id)
        return value

    def get_check_name_id(self, row: int, strict: bool = False) -> str:
        """Get check_name_id from work_sheet."""
        return self.get_goal_name_id(row, strict)

    def get_rule_name_id(self, row: int, strict: bool = False) -> str:
        """Get rule_name_id from work_sheet."""
        col = self._get_column_letter(self._column.rule_name_id)
        value = self._work_sheet[col + str(row)].value
        if value is None:
            self._add_row(row, self.rows_missing_rule_name_id)
        else:
            value = str(value).strip()
            if strict:
                svalue = str(value).strip()
                value = ''.join(str(svalue).split())
                if value != svalue:
                    self._add_row(row, self.rows_invalid_rule_name_id)
        return value

    def get_parameter_usage(self, row: int) -> str:
        """Get parameter_usage from work_sheet."""
        return self.get_goal_remarks(row)

    def get_parameter_value_default(self, row: int) -> str:
        """Get parameter_value_default from work_sheet."""
        col = self._get_column_letter(self._column.rename_values_alternatives)
        value = self._work_sheet[col + str(row)].value
        if value is not None:
            value = str(value).split(',')[0].strip()
        return value

    def get_parameter_values(self, row: int) -> str:
        """Get parameter_values from work_sheet."""
        col = self._get_column_letter(self._column.rename_values_alternatives)
        value = self._work_sheet[col + str(row)].value
        if value is None and self.get_parameter_name(row) is not None:
            self._add_row(row, self.rows_missing_parameters_values)
        # massage into comma separated list of values
        else:
            value = str(value).strip().replace(' ', '')
            value = value.replace(',[]', '')
            value = value.replace('[', '')
            value = value.replace(']', '')
            value = value.split(',')
        return value

    def _get_goal_text(self, row: int) -> str:
        """Get goal_text from work_sheet."""
        col = self._get_column_letter(self._column.control_text)
        goal_text = self._work_sheet[col + str(row)].value
        # normalize & tokenize
        value = goal_text.replace('\t', ' ')
        return value

    def _get_goal_text_tokens(self, row: int) -> List[str]:
        """Get goal_text tokens from work_sheet."""
        goal_text = self._get_goal_text(row)
        tokens = goal_text.split()
        return tokens

    def get_goal_remarks(self, row: int) -> str:
        """Get goal_remarks from work_sheet."""
        tokens = self._get_goal_text_tokens(row)
        # replace "Check whether" with "Ensure", if present
        if tokens:
            if tokens[0] == 'Check':
                if len(tokens) > 1:
                    if tokens[1] == 'whether':
                        tokens.pop(0)
                tokens[0] = 'Ensure'
        value = ' '.join(tokens)
        return value

    def get_controls(self, row: int) -> Dict[str, List[str]]:
        """Produce dict of controls mapped to statements.

        Example: {'au-2': ['(a)', '(d)'], 'au-12': [], 'si-4': ['(a)', '(b)', '(c)']}
        """
        value = {}
        for col in self._get_column_letter(self._column.nist_mappings):
            control = self._work_sheet[col + str(row)].value
            if control is None:
                continue
            # remove blanks
            control = ''.join(control.split())
            if len(control) < 1 or control.lower() == 'none':
                continue
            # remove rhs of : inclusive
            if ':' in control:
                control = control.split(':')[0]
            # remove alphabet parts of control & accumulate in statements
            control, statements = self._normalize_control(control)
            # skip bogus control made up if dashes only
            if len(control.replace('-', '')) == 0:
                continue
            if control not in value.keys():
                value[control] = statements
        if len(value.keys()) == 0:
            self._add_row(row, self.rows_missing_controls)
        logger.debug(f'row: {row} controls {value}')
        return value

    def get_component_name(self, row: int) -> str:
        """Get component_name from work_sheet."""
        col = self._get_column_letter(self._column.resource_title)
        value = self._work_sheet[col + str(row)].value
        if value is None:
            raise RuntimeError(f'row {row} col {col} missing component name')
        return value.strip()

    def get_parameter_name(self, row: int) -> Tuple[str, str]:
        """Get parameter_name from work_sheet."""
        return self.get_parameter_name_and_description(row)[0]

    def get_parameter_name_and_description(self, row: int) -> Tuple[str, str]:
        """Get parameter_name and description from work_sheet."""
        name = None
        description = None
        col = self._get_column_letter(self._column.rename_parameter_opt_parm)
        combined_values = self._work_sheet[col + str(row)].value
        if combined_values is not None:
            if '\n' in combined_values:
                parameter_parts = combined_values.split('\n')
            elif ' ' in combined_values:
                parameter_parts = combined_values.split(' ', 1)
            else:
                parameter_parts = combined_values
            if len(parameter_parts) == 2:
                name = parameter_parts[1].strip()
                description = parameter_parts[0].strip()
                sname = str(name).strip()
                name = sname.replace(' ', '_')
                if name != sname:
                    self._add_row(row, self.rows_invalid_parameter_name)
            else:
                logger.info(f'row {row} col {col} invalid value')
        if name is None and self.get_parameter_value_default(row) is not None:
            self._add_row(row, self.rows_missing_parameters)
        value = name, description
        return value

    def _get_control_id(self, row: int) -> int:
        """Get control_id from work_sheet."""
        col = self._get_column_letter(self._column.control_id)
        value = self._work_sheet[col + str(row)].value
        return value

    def _get_column_letter(self, name: str) -> str:
        """Get column letter."""
        value = self.map_name_to_letters[name]
        if len(value) == 1:
            value = value[0]
        return value

    def _map_columns(self) -> None:
        """Map columns."""
        self.map_name_to_letters = {}
        columns = self._work_sheet.max_column
        for column in range(1, columns + 1):
            cell_value = self._cell_value(1, column)
            if cell_value is None:
                continue
            cell_tokens = cell_value.split()
            normalized_cell_value = ' '.join(cell_tokens)
            # find columns of interest
            if self._column.control_id in cell_tokens:
                self._add_column(self._column.control_id, column, 1)
            elif self._column.control_text in cell_tokens:
                self._add_column(self._column.control_text, column, 1)
            elif self._column.goal_name_id in cell_tokens:
                self._add_column(self._column.goal_name_id, column, 1)
            elif self._column.goal_version in cell_tokens:
                self._add_column(self._column.goal_version, column, 1)
            elif self._column.rule_name_id in cell_tokens:
                self._add_column(self._column.rule_name_id, column, 1)
            elif self._column.rule_version in cell_tokens:
                self._add_column(self._column.rule_version, column, 1)
            # parameters and alternatives (exact tokens match)
            elif cell_tokens == self._column.tokens_parameter_opt_parm:
                self._add_column(self._column.rename_parameter_opt_parm, column, 1)
            elif cell_tokens == self._column.tokens_values_alternatives:
                self._add_column(self._column.rename_values_alternatives, column, 1)
            # filter column (exact string match)
            elif self._column.filter_column == normalized_cell_value:
                self._add_column(self._column.filter_column, column, 1)
            # nist mappings and resource title (multiple columns match)
            elif is_ordered_sublist(self._column.tokens_nist_mappings, cell_tokens):
                self._add_column(self._column.nist_mappings, column, 0)
            elif self._column.resource_title in cell_tokens:
                self._add_column(self._column.resource_title, column, 0)
        # insure expected columns found
        for name in [self._column.control_id,
                     self._column.control_text,
                     self._column.rule_name_id,
                     self._column.rule_version,
                     self._column.goal_name_id,
                     self._column.goal_version,
                     self._column.nist_mappings,
                     self._column.resource_title,
                     self._column.rename_parameter_opt_parm,
                     self._column.rename_values_alternatives]:
            if name not in self.map_name_to_letters.keys():
                raise RuntimeError(f'missing column {name}')

    def _add_column(self, name: str, column: int, limit: int) -> None:
        """Add column."""
        if name not in self.map_name_to_letters:
            self.map_name_to_letters[name] = []
        if limit > 0 and len(self.map_name_to_letters[name]) == limit:
            raise RuntimeError(f'duplicate column {name} {get_column_letter(column)}')
        self.map_name_to_letters[name].append(get_column_letter(column))

    def _cell_value(self, row: int, col: int) -> Any:
        """Get value for cell, adjusting for merged cells."""
        cell = self._work_sheet.cell(row, col)
        retval = cell.value
        if isinstance(cell, MergedCell):
            # cell is merged
            for mc_range in self._work_sheet.merged_cells.ranges:
                coord = get_column_letter(col) + str(row)
                if coord in mc_range:
                    retval = mc_range.start_cell.value
        return retval

    def _normalize_control(self, control: str) -> Tuple[str, List[str]]:
        """Remove parenthesized characters from controls."""
        statements = []
        for i in string.ascii_lowercase:
            needle = '(' + i + ')'
            if needle in control:
                statements.append(needle)
                control = control.replace(needle, '')
        control = control.lower()
        return control, statements

    def _add_row(self, row: int, account: List[int]) -> None:
        """Add row to accounting list of rows."""
        if row not in account:
            account.append(row)

    def report_issues(self) -> None:
        """Report issues."""
        if self.rows_missing_control_id:
            logger.info(f'rows missing control_id: {self.rows_missing_control_id}')
        if self.rows_invalid_goal_name_id:
            logger.info(f'rows invalid goal_name_id: {self.rows_invalid_goal_name_id}')
        if self.rows_missing_rule_name_id:
            logger.info(f'rows missing rule_name_id: {self.rows_missing_rule_name_id}')
        if self.rows_invalid_rule_name_id:
            logger.info(f'rows invalid rule_name_id: {self.rows_invalid_rule_name_id}')
        if self.rows_invalid_parameter_name:
            logger.info(f'rows invalid parameter_name: {self.rows_invalid_parameter_name}')
        if self.rows_missing_controls:
            logger.info(f'rows missing controls: {self.rows_missing_controls}')
        if self.rows_missing_parameters:
            logger.info(f'rows missing parameters: {self.rows_missing_parameters}')
        if self.rows_missing_parameters_values:
            logger.info(f'rows missing parameters values: {self.rows_missing_parameters_values}')
        if self.rows_filtered:
            logger.info(f'rows filtered: {self.rows_filtered}')
