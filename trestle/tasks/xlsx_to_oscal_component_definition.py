# -*- mode:python; coding:utf-8 -*-
# Copyright (c) 2021 IBM Corp. All rights reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     https://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
"""OSCAL transformation tasks."""

import configparser
import datetime
import logging
import pathlib
import string
import traceback
import uuid
from typing import Any, Dict, Iterator, List, Optional

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from trestle import __version__
from trestle.oscal import OSCAL_VERSION
from trestle.oscal.catalog import Catalog
from trestle.oscal.common import Link
from trestle.oscal.common import Metadata
from trestle.oscal.common import Parameter
from trestle.oscal.common import ParameterGuideline
from trestle.oscal.common import ParameterValue
from trestle.oscal.common import Party
from trestle.oscal.common import Property
from trestle.oscal.common import Remarks
from trestle.oscal.common import ResponsibleParty
from trestle.oscal.common import ResponsibleRole
from trestle.oscal.common import Role
from trestle.oscal.component import ComponentDefinition
from trestle.oscal.component import ControlImplementation
from trestle.oscal.component import DefinedComponent
from trestle.oscal.component import ImplementedRequirement
from trestle.oscal.component import SetParameter
from trestle.oscal.component import Statement
from trestle.tasks.base_task import TaskBase
from trestle.tasks.base_task import TaskOutcome
from trestle.utils.oscal_helper import CatalogHelper

logger = logging.getLogger(__name__)

t_component_name = str
t_control = str
t_description = str
t_goal_id = int
t_goal_name_id = str
t_goal_remarks = str
t_goal_text = str
t_goal_version = str
t_guidelines = str
t_name = str
t_parameter_usage = str
t_parameter_value = str
t_parameter_values = str
t_row = int
t_statements = List[str]
t_tokens = List[str]
t_type = str
t_uuid_str = str
t_values = str
t_work_sheet = Worksheet

t_controls = Dict[t_control, t_statements]


def get_trestle_version():
    """Get trestle version wrapper."""
    return __version__


class XlsxToOscalComponentDefinition(TaskBase):
    """
    Task to create OSCAL ComponentDefinition json.

    Attributes:
        name: Name of the task.
    """

    name = 'xlsx-to-oscal-component-definition'

    def __init__(self, config_object: Optional[configparser.SectionProxy]) -> None:
        """
        Initialize trestle task xlsx-to-oscal-component-definition.

        Args:
            config_object: Config section associated with the task.
        """
        super().__init__(config_object)
        self._timestamp = datetime.datetime.utcnow().replace(microsecond=0).replace(tzinfo=datetime.timezone.utc
                                                                                    ).isoformat()

    def set_timestamp(self, timestamp) -> None:
        """Set the timestamp."""
        self._timestamp = timestamp

    def print_info(self) -> None:
        """Print the help string."""
        logger.info(f'Help information for {self.name} task.')
        logger.info('')
        logger.info('Purpose: From spread sheet and catalog produce OSCAL component definition file.')
        logger.info('')
        logger.info('Configuration flags sit under [task.xlsx-to-oscal-component-definition]:')
        text1 = '  catalog-file      = '
        text2 = '(required) the path of the OSCAL catalog file.'
        logger.info(text1 + text2)
        text1 = '  spread-sheet-file = '
        text2 = '(required) the path of the spread sheet file.'
        logger.info(text1 + text2)
        text1 = '  work-sheet-name   = '
        text2 = '(required) the name of the work sheet in the spread sheet file.'
        logger.info(text1 + text2)
        text1 = '                      '
        text2 = 'column "ControlId" contains goal ID.'
        logger.info(text1 + text2)
        text2 = 'column "ControlText" contains goal text.'
        logger.info(text1 + text2)
        text2 = 'columns "NIST Mappings" contain controls.'
        logger.info(text1 + text2)
        text2 = 'column "ResourceTitle" contains component name.'
        logger.info(text1 + text2)
        text2 = 'column "goal_name_id" contains goal name.'
        logger.info(text1 + text2)
        text2 = 'column "Parameter [optional parameter]" contains parameter name + description, separated by newline.'
        logger.info(text1 + text2)
        text2 = 'column "Values [alternatives]" contains parameter values.'
        logger.info(text1 + text2)
        text1 = '  output-dir        = '
        text2 = '(required) the path of the output directory for synthesized OSCAL .json files.'
        logger.info(text1 + text2)
        text1 = '  output-overwrite  = '
        text2 = '(optional) true [default] or false; replace existing output when true.'
        logger.info(text1 + text2)

    def simulate(self) -> TaskOutcome:
        """Provide a simulated outcome."""
        return TaskOutcome('simulated-success')

    def execute(self) -> TaskOutcome:
        """Provide an executed outcome."""
        try:
            return self._execute()
        except Exception:
            logger.info(traceback.format_exc())
            return TaskOutcome('failure')

    def _execute(self) -> TaskOutcome:
        if not self._config:
            logger.error('config missing')
            return TaskOutcome('failure')
        # process config
        catalog_file = self._config.get('catalog-file')
        if catalog_file is None:
            logger.error('config missing "catalog-file"')
            return TaskOutcome('failure')
        self.catalog_helper = CatalogHelper(catalog_file)
        if not self.catalog_helper.exists():
            logger.error('"catalog-file" not found')
            return TaskOutcome('failure')
        spread_sheet = self._config.get('spread-sheet-file')
        if spread_sheet is None:
            logger.error('config missing "spread-sheet"')
            return TaskOutcome('failure')
        if not pathlib.Path(spread_sheet).exists():
            logger.error('"spread-sheet" not found')
            return TaskOutcome('failure')
        odir = self._config.get('output-dir')
        opth = pathlib.Path(odir)
        self._overwrite = self._config.getboolean('output-overwrite', True)
        quiet = self._config.get('quiet', False)
        self._verbose = not quiet
        # insure output dir exists
        opth.mkdir(exist_ok=True, parents=True)
        # announce spreadsheet
        if self._verbose:
            logger.info(f'input: {spread_sheet}')
        # calculate output file name & check writability
        oname = 'component-definition.json'
        ofile = opth / oname
        if not self._overwrite and pathlib.Path(ofile).exists():
            logger.error(f'output: {ofile} already exists')
            return TaskOutcome('failure')
        # initialize
        self.component_names = []
        self.defined_components = []
        self.parameters = {}
        self.parameter_helper = None
        # load the .xlsx contents
        wb = load_workbook(spread_sheet)
        sheet_name = self._config.get('work-sheet-name')
        if sheet_name is None:
            logger.error('config missing "work-sheet-name"')
            return TaskOutcome('failure')
        work_sheet = wb[sheet_name]
        # map column headings to column letters
        self._map_columns(work_sheet)
        # accumulators
        self.rows_missing_goal_name_id = []
        self.rows_missing_controls = []
        self.rows_missing_parameters = []
        self.rows_missing_parameters_values = []
        # roles, responsible_roles, parties, responsible parties
        party_uuid_01 = str(uuid.uuid4())
        party_uuid_02 = str(uuid.uuid4())
        party_uuid_03 = str(uuid.uuid4())
        roles = self._build_roles()
        responsible_roles = self._build_responsible_roles(party_uuid_01, party_uuid_02, party_uuid_03)
        parties = self._build_parties(party_uuid_01, party_uuid_02, party_uuid_03)
        responsible_parties = self._build_responsible_parties(party_uuid_01, party_uuid_02, party_uuid_03)
        # process each row of spread sheet
        for row in self._row_generator(work_sheet):
            # quit when first row with no goal_id encountered
            goal_name_id = self._get_goal_name_id(work_sheet, row)
            controls = self._get_controls(work_sheet, row)
            if len(controls.keys()) == 0:
                continue
            # component
            component_name = self._get_component_name(work_sheet, row)
            defined_component = self._get_defined_component(component_name)
            # parameter
            parameter_name, parameter_description = self._get_parameter_name_and_description(work_sheet, row)
            self._add_parameter(row, work_sheet, component_name, parameter_name, parameter_description)
            # implemented requirements
            self.implemented_requirements = []
            self._add_implemented_requirements(
                row, work_sheet, controls, component_name, parameter_name, responsible_roles, goal_name_id
            )
            # control implementations
            control_implementation = ControlImplementation(
                uuid=str(uuid.uuid4()),
                source=self._get_catalog_url(),
                description=component_name + ' implemented controls for ' + self._get_catalog_title()
                + '. It includes assessment asset configuration for CICD."',
                implemented_requirements=self.implemented_requirements,
            )
            if defined_component.control_implementations is None:
                defined_component.control_implementations = [control_implementation]
            else:
                defined_component.control_implementations.append(control_implementation)
        # create OSCAL ComponentDefinition
        metadata = Metadata(
            title='Component definition for ' + self._get_catalog_title() + ' profiles',
            last_modified=self._timestamp,
            oscal_version=OSCAL_VERSION,
            version=get_trestle_version(),
            roles=roles,
            parties=parties,
            responsible_parties=responsible_parties
        )
        component_definition = ComponentDefinition(
            uuid=str(uuid.uuid4()),
            metadata=metadata,
            components=self.defined_components,
        )
        # write OSCAL ComponentDefinition to file
        if self._verbose:
            logger.info(f'output: {ofile}')
        component_definition.oscal_write(pathlib.Path(ofile))
        # issues
        self._report_issues()
        # <hack>
        # create a catalog containing the parameters,
        # since parameters are not supported in OSCAL 1.0.0 component definition
        self._write_catalog()
        # </hack>
        return TaskOutcome('success')

    def _map_columns(self, work_sheet) -> None:
        """Map columns."""
        self.map_name_to_letters = {}
        columns = work_sheet.max_column
        for column in range(1, columns + 1):
            cell_value = self._cell_value(work_sheet, 1, column)
            # equal
            if self._fuzzy_equal('ControlId', cell_value):
                self._add_column('ControlId', column, 1)
            elif self._fuzzy_equal('ControlText', cell_value):
                self._add_column('ControlText', column, 1)
            elif self._fuzzy_equal('Version', cell_value):
                self._add_column('Version', column, 1)
            elif self._fuzzy_equal('goal_name_id', cell_value):
                self._add_column('goal_name_id', column, 1)
            # in
            elif self._fuzzy_in('NIST Mappings', cell_value):
                self._add_column('NIST Mappings', column, 0)
            elif self._fuzzy_in('ResourceTitle', cell_value):
                self._add_column('ResourceTitle', column, 1)
            # in and in
            if self._fuzzy_in('Parameter', cell_value) and self._fuzzy_in('[optional parameter]', cell_value):
                self._add_column('ParameterName', column, 1)
            if self._fuzzy_in('Values', cell_value) and self._fuzzy_in('[alternatives]', cell_value):
                self._add_column('ParameterValues', column, 1)
        for name in ['ControlId',
                     'ControlText',
                     'Version',
                     'goal_name_id',
                     'NIST Mappings',
                     'ResourceTitle',
                     'ParameterName',
                     'ParameterValues']:
            if name not in self.map_name_to_letters.keys():
                raise RuntimeError(f'missing column {name}')

    def _add_column(self, name, column, limit):
        """Add column."""
        if name not in self.map_name_to_letters:
            self.map_name_to_letters[name] = []
        if limit > 0:
            if len(self.map_name_to_letters[name]) == limit:
                raise RuntimeError(f'duplicate column {name} {get_column_letter(column)}')
        self.map_name_to_letters[name].append(get_column_letter(column))

    def _get_column_letter(self, name):
        """Get column letter."""
        value = self.map_name_to_letters[name]
        if len(value) == 1:
            value = value[0]
        return value

    def _fuzzy_equal(self, v1, v2) -> bool:
        """Fuzzy equal."""
        if v1 is None or v2 is None:
            value = False
        else:
            value = v1.lower().replace(' ', '') == v2.lower().replace(' ', '')
        return value

    def _fuzzy_in(self, v1, v2) -> bool:
        """Fuzzy in."""
        if v1 is None or v2 is None:
            value = False
        else:
            value = v1.lower().replace(' ', '') in v2.lower().replace(' ', '')
        return value

    def _cell_value(self, work_sheet, row, col) -> Any:
        """Get value for cell, adjusting for merged cells."""
        cell = work_sheet.cell(row, col)
        retval = cell.value
        if isinstance(cell, MergedCell):
            # cell is merged
            for mc_range in work_sheet.merged_cells.ranges:
                coord = get_column_letter(col) + str(row)
                if coord in mc_range:
                    retval = mc_range.start_cell.value
        return retval

    def _add_implemented_requirements(
        self, row, work_sheet, controls, component_name, parameter_name, responsible_roles, goal_name_id
    ) -> None:
        """Add implemented requirements."""
        goal_remarks = self._get_goal_remarks(work_sheet, row)
        parameter_value_default = self._get_parameter_value_default(work_sheet, row)
        for control in controls.keys():
            control_uuid = str(uuid.uuid4())
            prop1 = Property(
                name='goal_name_id',
                class_=self._get_class_for_property_name('goal_name_id'),
                value=goal_name_id,
                ns=self._get_namespace(),
                remarks=Remarks(__root__=str(goal_remarks))
            )
            prop2 = Property(
                name='goal_version',
                class_=self._get_class_for_property_name('goal_version'),
                value=self._get_goal_version(),
                ns=self._get_namespace(),
                remarks=Remarks(__root__=str(goal_name_id))
            )
            props = [prop1, prop2]
            control_id, status = self.catalog_helper.find_control_id(control)
            if control_id is None:
                logger.info(f'row {row} control {control} not found in catalog')
                control_id = control
            # implemented_requirement
            implemented_requirement = ImplementedRequirement(
                uuid=control_uuid,
                description=control,
                props=props,
                control_id=control_id,
                responsible_roles=responsible_roles,
            )
            # add statements
            self._add_statements(row, control, controls, component_name, implemented_requirement)
            # add set_parameter
            self._add_set_parameter(row, parameter_name, parameter_value_default, implemented_requirement)
            # implemented_requirements
            self.implemented_requirements.append(implemented_requirement)

    def _add_parameter(self, row, work_sheet, component_name, parameter_name, parameter_description) -> None:
        """Add_parameter."""
        usage = self._get_parameter_usage(work_sheet, row)
        if parameter_name is not None:
            parameter_name = parameter_name.strip()
            if ' ' in parameter_name:
                parameter_name = parameter_name.replace(' ', '_')
                logger.info(f'row={row} edited {parameter_name} to remove whitespace')
            values = self._get_parameter_values(work_sheet, row)
            guidelines = self._get_guidelines(values)
            href = self._get_namespace() + '/' + component_name.replace(' ', '%20')
            self.parameter_helper = ParameterHelper(
                values=values,
                id_=parameter_name,
                label=parameter_description,
                href=href,
                usage=usage,
                guidelines=guidelines,
            )
            self.parameters[str(uuid.uuid4())] = self.parameter_helper.get_parameter()

    def _add_statements(self, row, control, controls, component_name, implemented_requirement) -> None:
        """Add statements."""
        control_statements = controls[control]
        if len(control_statements) > 0:
            statements = []
            for control_statement in control_statements:
                statement_id = control + control_statement
                if any(i in control for i in '()'):
                    control = control.replace('(', '_')
                    control = control.replace(')', '')
                    logger.info(f'row {row} control {control} edited to remove parentheses')
                statement = Statement(
                    statement_id=control,
                    uuid=str(uuid.uuid4()),
                    description=f'{component_name} implements {statement_id}'
                )
                statements.append(statement)
            implemented_requirement.statements = statements

    def _add_set_parameter(self, row, parameter_name, parameter_value_default, implemented_requirement) -> None:
        """Add set parameter."""
        if parameter_name is None:
            if row not in self.rows_missing_parameters:
                self.rows_missing_parameters.append(row)
        else:
            parameter_name = parameter_name.replace(' ', '_')
            if parameter_value_default is None:
                if row not in self.rows_missing_parameters_values:
                    self.rows_missing_parameters_values.append(row)
            else:
                values = [parameter_value_default]
                set_parameter = SetParameter(param_id=parameter_name, values=values)
                set_parameters = [set_parameter]
                implemented_requirement.set_parameters = set_parameters

    def _get_defined_component(self, component_name):
        """Get defined component."""
        if component_name not in self.component_names:
            # create new component
            self.component_names.append(component_name)
            component_title = component_name
            component_description = component_name
            defined_component = DefinedComponent(
                uuid=str(uuid.uuid4()),
                description=component_description,
                title=component_title,
                type='Service',
            )
            self.defined_components.append(defined_component)
        else:
            # find existing component
            for defined_component in self.defined_components:
                if component_name == defined_component.title:
                    break
        return defined_component

    def _build_roles(self):
        """Build roles."""
        value = [
            Role(id='prepared-by', title='Indicates the organization that created this content.'),
            Role(id='prepared-for', title='Indicates the organization for which this content was created..'),
            Role(
                id='content-approver',
                title='Indicates the organization responsible for all content represented in the "document".'
            ),
        ]
        return value

    def _build_responsible_roles(self, party_uuid_01, party_uuid_02, party_uuid_03):
        """Build responsible roles."""
        role_prepared_by = ResponsibleRole(role_id='prepared-by', party_uuids=[party_uuid_01])
        role_prepared_for = ResponsibleRole(role_id='prepared-for', party_uuids=[party_uuid_02, party_uuid_03])
        role_content_approver = ResponsibleRole(role_id='content-approver', party_uuids=[party_uuid_01])
        value = [
            role_prepared_by,
            role_prepared_for,
            role_content_approver,
        ]
        return value

    def _build_parties(self, party_uuid_01, party_uuid_02, party_uuid_03):
        """Build parties."""
        value = [
            Party(uuid=party_uuid_01, type='organization', name=self._get_org_name(), remarks=self._get_org_remarks()),
            Party(
                uuid=party_uuid_02,
                type='organization',
                name='Customer',
                remarks='organization to be customized at account creation only for their Component Definition'
            ),
            Party(
                uuid=party_uuid_03,
                type='organization',
                name='ISV',
                remarks='organization to be customized at ISV subscription only for their Component Definition'
            ),
        ]
        return value

    def _build_responsible_parties(self, party_uuid_01, party_uuid_02, party_uuid_03):
        """Build responsible parties."""
        prepared_by = ResponsibleParty(role_id='prepared-by', party_uuids=[party_uuid_01])
        prepared_for = ResponsibleParty(role_id='prepared-for', party_uuids=[party_uuid_02, party_uuid_03])
        content_approver = ResponsibleParty(role_id='content-approver', party_uuids=[party_uuid_01])
        value = [
            prepared_by,
            prepared_for,
            content_approver,
        ]
        return value

    def _report_issues(self):
        """Report issues."""
        if len(self.rows_missing_goal_name_id) > 0:
            logger.info(f'rows missing goal_name_id: {self.rows_missing_goal_name_id}')
        if len(self.rows_missing_controls) > 0:
            logger.info(f'rows missing controls: {self.rows_missing_controls}')
        if len(self.rows_missing_parameters) > 0:
            logger.info(f'rows missing parameters: {self.rows_missing_parameters}')
        if len(self.rows_missing_parameters_values) > 0:
            logger.info(f'rows missing parameters values: {self.rows_missing_parameters_values}')

    def _write_catalog(self):
        """Create a catalog containing the parameters."""
        if self.parameter_helper is not None:
            tdir = self._config.get('output-dir')
            tdir = tdir.replace('component-definitions', 'catalogs')
            tpth = pathlib.Path(tdir)
            tname = 'catalog.json'
            tfile = tpth / tname
            self.parameter_helper.write_parameters_catalog(
                parameters=self.parameters,
                timestamp=self._timestamp,
                oscal_version=OSCAL_VERSION,
                version=get_trestle_version(),
                ofile=tfile,
                verbose=self._verbose,
            )

    def _get_org_name(self) -> str:
        """Get org-name from config."""
        value = self._config.get('org-name')
        logger.debug(f'org-name: {value}')
        return value

    def _get_org_remarks(self) -> str:
        """Get org-remarks from config."""
        value = self._config.get('org-remarks')
        logger.debug(f'org-remarks: {value}')
        return value

    def _get_class_for_property_name(self, property_name) -> str:
        """Get class for property-name from config."""
        value = None
        data = self._config.get('property-name-to-class')
        if data is not None:
            for item in data.split(','):
                item = item.strip()
                parts = item.split(':')
                if len(parts) == 2:
                    if parts[0] == property_name:
                        value = parts[1]
                        break
        logger.debug(f'property-name-to-class: {property_name} -> {value}')
        return value

    def _get_namespace(self) -> str:
        """Get namespace from config."""
        value = self._config.get('namespace')
        logger.debug(f'namespace: {value}')
        return value

    def _get_catalog_url(self) -> str:
        """Get catalog url from config."""
        value = self._config.get('catalog-url')
        logger.debug(f'catalog-url: {value}')
        return value

    def _get_catalog_title(self) -> str:
        """Get catalog title from config."""
        value = self._config.get('catalog-title')
        logger.debug(f'catalog-title: {value}')
        return value

    def _row_generator(self, work_sheet: t_work_sheet) -> Iterator[t_row]:
        """Generate rows until goal_id is None."""
        row = 1
        while True:
            row = row + 1
            goal_id = self._get_goal_id(work_sheet, row)
            if goal_id is None:
                break
            yield row

    def _get_goal_version(self) -> t_goal_version:
        """Fix goal_version at 1.0."""
        return '1.0'

    def _get_goal_id(self, work_sheet: t_work_sheet, row: t_row) -> t_goal_id:
        """Get goal_id from work_sheet."""
        col = self._get_column_letter('ControlId')
        value = work_sheet[col + str(row)].value
        return value

    def _get_goal_text(self, work_sheet: t_work_sheet, row: t_row) -> t_goal_text:
        """Get goal_text from work_sheet."""
        col = self._get_column_letter('ControlText')
        goal_text = work_sheet[col + str(row)].value
        # normalize & tokenize
        value = goal_text.replace('\t', ' ')
        return value

    def _get_controls(self, work_sheet: t_work_sheet, row: t_row) -> t_controls:
        """Produce dict of controls mapped to statements.

        Example: {'au-2': ['(a)', '(d)'], 'au-12': [], 'si-4': ['(a)', '(b)', '(c)']}
        """
        value = {}
        for col in self._get_column_letter('NIST Mappings'):
            control = work_sheet[col + str(row)].value
            if control is not None:
                # remove blanks
                control = ''.join(control.split())
                if len(control) > 0:
                    if control.lower() == 'none':
                        continue
                    # remove rhs of : inclusive
                    if ':' in control:
                        control = control.split(':')[0]
                    # remove alphabet parts of control & accumulate in statements
                    statements = []
                    for i in string.ascii_lowercase:
                        needle = '(' + i + ')'
                        if needle in control:
                            statements.append(needle)
                            control = control.replace(needle, '')
                    control = control.lower()
                    # skip bogus control made up if dashes only
                    if len(control.replace('-', '')) == 0:
                        continue
                    if control not in value.keys():
                        value[control] = statements
        if len(value.keys()) == 0:
            self.rows_missing_controls.append(row)
        logger.debug(f'row: {row} controls {value}')
        return value

    def _get_goal_name_id(self, work_sheet: t_work_sheet, row: t_row) -> t_goal_name_id:
        """Get goal_name_id from work_sheet."""
        col = self._get_column_letter('goal_name_id')
        value = work_sheet[col + str(row)].value
        if value is None:
            self.rows_missing_goal_name_id.append(row)
            value = self._get_goal_id(work_sheet, row)
        value = str(value).strip()
        return value

    def _get_parameter_name_and_description(self, work_sheet: t_work_sheet, row: t_row) -> (t_name, t_description):
        """Get parameter_name and description from work_sheet."""
        name = None
        description = None
        col = self._get_column_letter('ParameterName')
        combined_values = work_sheet[col + str(row)].value
        if combined_values is not None:
            if '\n' in combined_values:
                parameter_parts = combined_values.split('\n')
            elif ' ' in combined_values:
                parameter_parts = combined_values.split(' ', 1)
            else:
                parameter_parts = combined_values
            if len(parameter_parts) != 2:
                raise RuntimeError(f'row {row} col {col} unable to parse')
            name = parameter_parts[1].strip()
            description = parameter_parts[0].strip()
        value = name, description
        return value

    def _get_parameter_value_default(self, work_sheet: t_work_sheet, row: t_row) -> t_parameter_value:
        """Get parameter_value_default from work_sheet."""
        col = self._get_column_letter('ParameterValues')
        value = work_sheet[col + str(row)].value
        if value is not None:
            value = str(value).split(',')[0].strip()
        return value

    def _get_parameter_values(self, work_sheet: t_work_sheet, row: t_row) -> t_parameter_values:
        """Get parameter_values from work_sheet."""
        col = self._get_column_letter('ParameterValues')
        value = work_sheet[col + str(row)].value
        if value is None:
            logger.info(f'row {row} col {col} missing value')
        # massage into comma separated list of values
        value = str(value).strip().replace(' ', '')
        value = value.replace(',[]', '')
        value = value.replace('[', '')
        value = value.replace(']', '')
        return value

    def _get_guidelines(self, values: t_values) -> t_guidelines:
        """Get guidelines based on values."""
        type_ = self._get_type(values)
        value = 'The first listed value option is set by default in the system '
        value += 'unless set-parameter is used to satisfy a control requirements. '
        value += f'Type {type_}.'
        return value

    def _get_type(self, values: t_values) -> t_type:
        """Get type based on values."""
        if self._is_int(values):
            value = 'Integer'
        elif self._is_float(values):
            value = 'Float'
        else:
            value = 'String'
        return value

    def _is_int(self, values: t_values) -> bool:
        """Determine if string represents list of int."""
        try:
            for value in values.split(','):
                int(value)
            retval = True
        except Exception:
            retval = False
        return retval

    def _is_float(self, values: t_values) -> bool:
        """Determine if string represents list of float."""
        try:
            for value in values.split(','):
                float(value)
            retval = True
        except Exception:
            retval = False
        return retval

    def _get_parameter_usage(self, work_sheet: t_work_sheet, row: t_row) -> t_parameter_usage:
        """Get parameter_usage from work_sheet."""
        return self._get_goal_remarks(work_sheet, row)

    def _get_goal_text_tokens(self, work_sheet: t_work_sheet, row: t_row) -> t_tokens:
        """Get goal_text tokens from work_sheet."""
        goal_text = self._get_goal_text(work_sheet, row)
        tokens = goal_text.split()
        return tokens

    def _get_goal_remarks(self, work_sheet: t_work_sheet, row: t_row) -> t_goal_remarks:
        """Get goal_remarks from work_sheet."""
        tokens = self._get_goal_text_tokens(work_sheet, row)
        # replace "Check whether" with "Ensure", if present
        if len(tokens) > 0:
            if tokens[0] == 'Check':
                if len(tokens) > 1:
                    if tokens[1] == 'whether':
                        tokens.pop(0)
                tokens[0] = 'Ensure'
        value = ' '.join(tokens)
        return value

    def _get_component_name(self, work_sheet: t_work_sheet, row: t_row) -> t_component_name:
        """Get component_name from work_sheet."""
        col = self._get_column_letter('ResourceTitle')
        value = work_sheet[col + str(row)].value
        if value is None:
            raise RuntimeError(f'row {row} col {col} missing component name')
        return value


t_guidelines = str
t_href = str
t_id = str
t_label = str
t_ofile = str
t_oscal_version = str
t_parameters = Dict[str, Parameter]
t_timestamp = str
t_usage = str
t_values = Any
t_verbose = bool
t_version = str


class ParameterHelper():
    """Parameter Helper class is a temporary hack because Component Definition does not support Parameters."""

    def __init__(
        self, values: t_values, id_: t_id, label: t_label, href: t_href, usage: t_usage, guidelines: t_guidelines
    ) -> None:
        """Initialize."""
        self._parameter_values = ParameterValue(__root__=str(values))
        self._id = id_
        self._label = label
        self._links = [Link(href=href)]
        self._usage = usage
        self._guidelines = ParameterGuideline(prose=guidelines)

    def get_parameter(self) -> Parameter:
        """Get parameter."""
        parameter = Parameter(
            id=self._id,
            label=self._label,
            links=self._links,
            usage=self._usage,
            guidelines=[self._guidelines],
            values=[self._parameter_values]
        )
        return parameter

    def write_parameters_catalog(
        self,
        parameters: t_parameters,
        timestamp: t_timestamp,
        oscal_version: t_oscal_version,
        version: t_version,
        ofile: t_ofile,
        verbose: t_verbose,
    ) -> None:
        """Write parameters catalog."""
        parameter_metadata = Metadata(
            title='Component Parameters',
            last_modified=timestamp,
            oscal_version=oscal_version,
            version=version,
        )
        parameter_catalog = Catalog(
            uuid=str(uuid.uuid4()),
            metadata=parameter_metadata,
            params=list(parameters.values()),
        )
        if verbose:
            logger.info(f'output: {ofile}')
        parameter_catalog.oscal_write(pathlib.Path(ofile))
