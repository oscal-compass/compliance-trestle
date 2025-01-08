# Copyright (c) 2025 The OSCAL Compass Authors. All rights reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     https://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
"""OSCAL transformation tasks."""
import configparser
import csv
import datetime
import logging
import pathlib
import tempfile
import traceback
from configparser import SectionProxy
from functools import cmp_to_key
from typing import Dict, Iterator, List, Optional

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from trestle.common.str_utils import as_bool
from trestle.tasks.base_task import TaskBase
from trestle.tasks.base_task import TaskOutcome
from trestle.tasks.csv_to_oscal_cd import CsvToOscalComponentDefinition

logger = logging.getLogger(__name__)

timestamp = datetime.datetime.now(datetime.timezone.utc).replace(microsecond=0).isoformat()

default_benchmark_control_prefix = 'cisc-'
default_benchmark_rule_prefix = 'CIS-'
output_file = 'component-definition.json'

head_recommendation_no = 'Recommendation #'
head_section_no = 'Section #'


class CisXlsxToOscalCd(TaskBase):
    """
    Task to transform CIS .xlsx to OSCAL component definition.

    Attributes:
        name: Name of the task.
    """

    name = 'cis-xlsx-to-oscal-cd'

    def __init__(self, config_object: Optional[configparser.SectionProxy]) -> None:
        """
        Initialize trestle task.

        Args:
            config_object: Config section associated with the task.
        """
        self.config_object = config_object
        super().__init__(config_object)
        self._default_benchmark_sheet_name = 'Combined Profiles'
        self._default_component_type = 'software'
        self._default_output_overwrite = 'True'
        #
        self._example_benchmark_file = 'data/CIS_IBM_Db2_11_Benchmark_v1.1.0.xlsx'
        self._example_benchmark_title = 'CIS IBM Db2 11 Benchmark'
        self._example_benchmark_version = '1.1.0'
        self._example_oscal_cd_dir = 'data/component-definitions/CIS_IBM_Db2_11_Benchmark_v1.1.0'
        self._example_namespace = 'https://oscal-compass/compliance-trestle/schemas/oscal/cd'
        self._example_profile_version = 'v8'
        self._example_profile_source = 'catalogs/CIS_controls_v8/catalog.json'
        self._example_profile_description = 'CIS catalog v8'
        self._example_component_name = 'Db2 11'

    def print_info(self) -> None:
        """Print the help string."""
        logger.info(f'Help information for {self.name} task.')
        logger.info('')
        logger.info('Purpose: Create component definition from standard CIS benchmark')
        logger.info('')
        logger.info('Configuration flags sit under [task.cis-xlsx-to-oscal-cd]:')
        text1 = '  benchmark-file             = '
        text2 = f'(required) path of file to read the CIS benchmark .xlsx, e.g., "{self._example_benchmark_file}".'
        logger.info(text1 + text2)
        text1 = '  benchmark-title            = '
        text2 = f'(required) title of the CIS benchmark, e.g., "{self._example_benchmark_title}".'
        logger.info(text1 + text2)
        text1 = '  benchmark-version          = '
        text2 = f'(required) version of the CIS benchmark .xlsx, e.g., "{self._example_benchmark_version}".'
        logger.info(text1 + text2)
        text1 = '  benchmark-control-prefix   = '
        text2 = f'(optional) benchmark control prefix, default = "{default_benchmark_control_prefix}".'
        logger.info(text1 + text2)
        text1 = '  benchmark-rule-prefix      = '
        text2 = f'(optional) benchmark rule prefix, default = "{default_benchmark_rule_prefix}".'
        logger.info(text1 + text2)
        text1 = '  benchmark-sheet-name       = '
        text2 = f'(optional) benchmark sheet name, default = "{self._default_benchmark_sheet_name}".'
        logger.info(text1 + text2)
        text1 = '  component-name             = '
        text2 = f'(required) component name, e.g., "{self._example_component_name}".'
        logger.info(text1 + text2)
        text1 = '  component-description      = '
        text2 = f'(required) component description, e.g., "{self._example_component_name}".'
        logger.info(text1 + text2)
        text1 = '  component-type             = '
        text2 = f'(required) component type, e.g., "{self._default_component_type}".'
        logger.info(text1 + text2)
        text1 = '  namespace                  = '
        text2 = f'(required) namespace, e.g., "{self._example_namespace}".'
        logger.info(text1 + text2)
        text1 = '  output-dir                 = '
        text2 = f'(required) path of folder to write the OSCAL {output_file}, e.g., "{self._example_oscal_cd_dir}".'
        logger.info(text1 + text2)
        text1 = '  output-overwrite           = '
        text2 = f'(optional) output overwrite, default = "{self._default_output_overwrite}".'
        logger.info(text1 + text2)
        text1 = '  profile-version            = '
        text2 = f'(required) profile version, e.g., "{self._example_profile_version}".'
        logger.info(text1 + text2)
        text1 = '  profile-source             = '
        text2 = f'(required) profile source, e.g., "{self._example_profile_source}".'
        logger.info(text1 + text2)
        text1 = '  profile-description        = '
        text2 = f'(required) profile description, e.g., "{self._example_profile_description}".'
        logger.info(text1 + text2)

    def simulate(self) -> TaskOutcome:
        """Provide a simulated outcome."""
        return TaskOutcome('simulated-success')

    def execute(self) -> TaskOutcome:
        """Provide an actual outcome."""
        try:
            return self._execute()
        except Exception:
            logger.info(traceback.format_exc())
            return TaskOutcome('failure')

    def _execute(self) -> TaskOutcome:
        """Wrap the execute for exception handling."""
        if not self.config_object:
            logger.warning('config missing')
            return TaskOutcome('failure')
        # required
        try:
            self._benchmark_file = self.config_object['benchmark-file']
            self._oscal_cd_dir = self.config_object['output-dir']
            self._namespace = self.config_object['namespace']
            self._profile_version = self.config_object['profile-version']
        except KeyError as e:
            logger.info(f'key {e.args[0]} missing')
            return TaskOutcome('failure')
        # output
        self._oscal_cd_path = pathlib.Path(self._oscal_cd_dir)
        # insure output dir exists
        self._oscal_cd_path.mkdir(exist_ok=True, parents=True)
        # calculate output file name & check writability
        oname = 'component-definition.json'
        ofile = self._oscal_cd_path / oname
        overwrite = self.config_object.get('output-overwrite', self._default_output_overwrite)
        overwrite = as_bool(overwrite)
        if not overwrite and pathlib.Path(ofile).exists():
            logger.warning(f'output: {ofile} already exists')
            return TaskOutcome('failure')
        with self._get_tempdir() as tmpdir:
            # step 1 - add combined sheet, if needed
            combine_helper = CombineHelper(self.config_object, tmpdir)
            combine_helper.run()
            # step 2 - create trestle ready csv file from xlsx file
            xlsx_to_csv_helper = XlsxToCsvHelper(self.config_object, tmpdir)
            xlsx_to_csv_helper.run()
            # step 3 - create OSCAL json file from csv file
            csv_to_json_helper = CsvToJsonHelper(self.config_object, tmpdir)
            task_outcome = csv_to_json_helper.run()
            return task_outcome

    def _get_tempdir(self) -> tempfile.TemporaryDirectory():
        """Get tmpdir."""
        return tempfile.TemporaryDirectory()


class SortHelper:
    """SortHelper."""

    @staticmethod
    def compare(item1: str, item2: str) -> int:
        """Compare."""
        # get parts
        parts1 = ''.split('.')
        if item1 is not None:
            parts1 = str(item1).split('.')
        parts2 = ''.split('.')
        if item2 is not None:
            parts2 = str(item2).split('.')
        # normalize parts length
        while len(parts1) < len(parts2):
            parts1.append('0')
        while len(parts2) < len(parts1):
            parts2.append('0')
        # comparison
        rval = 0
        for i in range(len(parts1)):
            try:
                v1 = int(parts1[i])
            except Exception:
                rval = -1
                break
            try:
                v2 = int(parts2[i])
            except Exception:
                rval = 1
                break
            if v1 < v2:
                rval = -1
                break
            if v1 > v2:
                rval = 1
                break
        text = f'compare rval: {rval} item1: {item1} item2: {item2}'
        logger.debug(f'{text}')
        return rval


class SheetHelper:
    """SheetHelper."""

    def __init__(self, wb: Workbook, sn: str) -> None:
        """Initialize."""
        self.wb = wb
        self.sn = sn
        self.ws = self.wb[self.sn]

    def get_sn(self) -> int:
        """Get sheet name."""
        return self.sn

    def get_max_col(self) -> int:
        """Get max column."""
        return self.ws.max_column

    def row_generator(self) -> Iterator[int]:
        """Generate rows until max reached."""
        row = 1
        while row <= self.ws.max_row:
            yield row
            row += 1

    def get_cell_value(self, row: int, col: int) -> str:
        """Get cell value for given row and column name."""
        cell = self.ws.cell(row, col)
        return cell.value

    def put_cell_value(self, row: int, col: int, value: str) -> None:
        """Get cell value for given row and column name."""
        cell = self.ws.cell(row, col)
        cell.value = value

    @staticmethod
    def get_sheetname_prefixes() -> List[str]:
        """Get sheetnames prefixes."""
        rval = ['Level 1', 'Level 2']
        return rval

    @staticmethod
    def get_sheetname() -> str:
        """Get sheetname output."""
        rval = 'Combined Profiles'
        return rval


class ColHelper:
    """Col Helper."""

    @staticmethod
    def get_section() -> int:
        """Get section col no."""
        return 1

    @staticmethod
    def get_recommendation() -> int:
        """Get recommendation col no."""
        return 2


class Int:
    """Int."""

    def __init__(self, value=0):
        """Initialize."""
        self.value = value

    def inc_value(self) -> None:
        """Increment."""
        self.value += 1

    def get_value(self) -> int:
        """Get."""
        return self.value


class CombineHelper:
    """Combine helper."""

    tgt_col_profile = 3

    def __init__(self, config: SectionProxy, tmpdir: str) -> None:
        """Initialize."""
        benchmark_file = config['benchmark-file']
        self.ipath = pathlib.Path(benchmark_file)
        self.opath = pathlib.Path(tmpdir) / self.ipath.name
        self.wb = load_workbook(self.ipath)
        self.ws_map = {}
        self.combined_map = {}

    def run(self) -> None:
        """Run."""
        self._add_sheet_combined_profiles()
        self._save()

    def _gather_sheets(self) -> None:
        """Gather sheets."""
        for sn in self.wb.sheetnames:
            for pn in self.sheetnames_prefixes:
                if sn.startswith(pn):
                    self.ws_map[sn] = SheetHelper(self.wb, sn)
                    logger.debug(f'input sheet {sn} to be combined.')
                    break

    def _validate_columns_count(self) -> None:
        """Validate columns count."""
        columns = -1
        for sn in self.ws_map.keys():
            sheet_helper = self.ws_map[sn]
            if columns < 0:
                columns = sheet_helper.get_max_col()
            if columns != sheet_helper.get_max_col():
                raise RuntimeError(f'{sn} unexpected columns count {sheet_helper.get_max_col()} for sheet {sn}')

    def _populate_combined_map(self) -> int:
        """Populate combined map."""
        src_col_section_no = ColHelper.get_section()
        src_col_recommendation_no = ColHelper.get_recommendation()
        rec_count_sheets = 0
        # populate combined map
        for sn in self.ws_map.keys():
            sheet_helper = SheetHelper(self.wb, sn)
            # process all rows from individual sheet
            rec_count_sheets += self._process_sheet(sheet_helper, src_col_section_no, src_col_recommendation_no)
        return rec_count_sheets

    def _process_sheet(self, sheet_helper: SheetHelper, src_col_section_no: int, src_col_recommendation_no: int) -> int:
        """Process sheet."""
        rec_count = 0
        for row in sheet_helper.row_generator():
            # section
            section_no = sheet_helper.get_cell_value(row, src_col_section_no)
            if section_no not in self.combined_map.keys():
                self.combined_map[section_no] = {}
            # recommendation
            recommendation_no = sheet_helper.get_cell_value(row, src_col_recommendation_no)
            if recommendation_no not in self.combined_map[section_no].keys():
                self.combined_map[section_no][recommendation_no] = {}
            # combine head or data
            if row == 1:
                self._combine_head(sheet_helper, row, section_no, recommendation_no, CombineHelper.tgt_col_profile)
            else:
                self._combine_data(sheet_helper, row, section_no, recommendation_no, CombineHelper.tgt_col_profile)
                if recommendation_no:
                    rec_count += 1
        return rec_count

    def _handle_head_row(
        self, combined_helper: SheetHelper, row: Int, kvset: Dict, section_no: str, recommendation_no: str
    ) -> None:
        """Handle head row."""
        for col in kvset.keys():
            value = self.combined_map[section_no][recommendation_no][col]
            if col == CombineHelper.tgt_col_profile:
                value = value[0]
            combined_helper.put_cell_value(row.get_value(), col, value)
        row.inc_value()

    def _handle_data_row_control(
        self,
        combined_helper: SheetHelper,
        row: Int,
        kvset: Dict,
        section_no: str,
        recommendation_no: str,
        rec_count_merged: Int
    ) -> None:
        """Handle data row control."""
        profiles = kvset[CombineHelper.tgt_col_profile]
        for profile in profiles:
            for col in kvset.keys():
                value = self.combined_map[section_no][recommendation_no][col]
                if col == CombineHelper.tgt_col_profile:
                    value = profile
                combined_helper.put_cell_value(row.get_value(), col, value)
            row.inc_value()
            rec_count_merged.inc_value()

    def _handle_data_row_non_control(
        self, combined_helper: SheetHelper, row: Int, kvset: Dict, section_no: str, recommendation_no: str
    ) -> None:
        """Handle data row non-control."""
        for col in kvset.keys():
            value = self.combined_map[section_no][recommendation_no][col]
            if col == CombineHelper.tgt_col_profile:
                value = None
            combined_helper.put_cell_value(row.get_value(), col, value)
        row.inc_value()

    def _handle_data_row(
        self,
        combined_helper: SheetHelper,
        row: Int,
        kvset: Dict,
        section_no: str,
        recommendation_no: str,
        rec_count_merged: Int
    ) -> None:
        """Handle data row."""
        if recommendation_no:
            self._handle_data_row_control(combined_helper, row, kvset, section_no, recommendation_no, rec_count_merged)
        else:
            self._handle_data_row_non_control(combined_helper, row, kvset, section_no, recommendation_no)

    def _populate_combined_sheet(self, combined_helper: SheetHelper) -> int:
        """Populate combined sheet."""
        rec_count_merged = Int(0)
        row = Int(1)
        keys1 = list(self.combined_map.keys())
        keys1.sort(key=cmp_to_key(SortHelper.compare))
        for section_no in keys1:
            section = self.combined_map[section_no]
            keys2 = list(section.keys())
            keys2.sort(key=cmp_to_key(SortHelper.compare))
            for recommendation_no in keys2:
                kvset = self.combined_map[section_no][recommendation_no]
                if row.get_value() == 1:
                    self._handle_head_row(combined_helper, row, kvset, section_no, recommendation_no)
                else:
                    self._handle_data_row(combined_helper, row, kvset, section_no, recommendation_no, rec_count_merged)
        return rec_count_merged.get_value()

    def _add_sheet_combined_profiles(self) -> None:
        """Add sheet combined profiles."""
        # output sheet
        self.sheetname_output = SheetHelper.get_sheetname()
        exists = self.sheetname_output in self.wb.sheetnames
        if exists:
            logger.debug(f'output sheet {self.sheetname_output} exists.')
            return
        # input sheets
        self.sheetnames_prefixes = SheetHelper.get_sheetname_prefixes()
        # sheets
        self._gather_sheets()
        # validate
        self._validate_columns_count()
        # key columns mappings
        rec_count_sheets = self._populate_combined_map()
        # add combined sheet
        sn = self.sheetname_output
        self.wb.create_sheet(sn)
        combined_helper = SheetHelper(self.wb, sn)
        self.ws_map[sn] = combined_helper
        # populate combined sheet
        rec_count_merged = self._populate_combined_sheet(combined_helper)
        # correctness check
        if rec_count_sheets != rec_count_merged:
            raise RuntimeError(f'recommendation counts original: {rec_count_sheets} merged: {rec_count_merged}')

    def _combine_head(
        self, sheet_helper: SheetHelper, row: int, section_no: str, recommendation_no: str, tgt_col_profile: int
    ) -> None:
        """Combine head."""
        if not len(self.combined_map[section_no][recommendation_no].keys()):
            self.combined_map[section_no][recommendation_no][tgt_col_profile] = ['profile']
            for col in range(1, sheet_helper.get_max_col() + 1):
                if col < tgt_col_profile:
                    tcol = col
                else:
                    tcol = col + 1
                value = sheet_helper.get_cell_value(row, col)
                self.combined_map[section_no][recommendation_no][tcol] = value

    def _combine_data(
        self, sheet_helper: SheetHelper, row: int, section_no: str, recommendation_no: str, tgt_col_profile: int
    ) -> None:
        """Combine data."""
        if not len(self.combined_map[section_no][recommendation_no].keys()):
            self.combined_map[section_no][recommendation_no][tgt_col_profile] = []
        sn = sheet_helper.get_sn()
        self.combined_map[section_no][recommendation_no][tgt_col_profile].append(sn)
        for col in range(1, sheet_helper.get_max_col() + 1):
            if col < tgt_col_profile:
                tcol = col
            else:
                tcol = col + 1
            self.combined_map[section_no][recommendation_no][tcol] = sheet_helper.get_cell_value(row, col)

    def _save(self) -> None:
        """Save."""
        self.wb.save(self.opath)
        logger.debug(f'{self.opath} saved')


class PropertyHelper:
    """Property Helper."""

    @staticmethod
    def normalize_name(name: str) -> str:
        """Normalize name."""
        rval = name.replace('(', '').replace(')', '').replace('#', '').strip()
        rval = rval.replace(' ', '_')
        rval = rval.replace('__', '_')
        return rval


class CsvHelper:
    """Csv Helper."""

    def __init__(self, path: pathlib.Path) -> None:
        """Initialize."""
        self.path = path
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.rows = []

    def add_row(self, row: List[str]) -> None:
        """Add row."""
        self.rows.append(row)

    def delete_last_row(self) -> None:
        """Delete last row."""
        self.rows = self.rows[:-1]

    def write(self) -> None:
        """Write csv file."""
        with open(self.path, 'w', newline='', encoding='utf-8') as output:
            csv_writer = csv.writer(output, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
            for row in self.rows:
                csv_writer.writerow(row)

    def save(self) -> None:
        """Save csv file."""
        self.write()
        logger.debug(f'{self.path} saved')

    @staticmethod
    def columns() -> Dict:
        """Columns."""
        return {
            'Component_Title': 'A human readable name for the component.',  # noqa
            'Component_Description': 'A description of the component including information about its function.',  # noqa
            'Component_Type': 'A category describing the purpose of the component. ALLOWED VALUES interconnection:software:hardware:service:physical:process-procedure:plan:guidance:standard:validation:',  # noqa
            'Profile': 'List of CIS profiles',  # noqa
            'Rule_Id': 'A textual label that uniquely identifies a policy (desired state) that can be used to reference it elsewhere in this or other documents.',  # noqa
            'Rule_Description': 'A description of the policy (desired state) including information about its purpose and scope.',  # noqa
            'Profile_Source': 'A URL reference to the source catalog or profile for which this component is implementing controls for. A profile designates a selection and configuration of controls from one or more catalogs.',  # noqa
            'Profile_Description': 'A description of the profile.',  # noqa
            'Control_Id_List': 'A list of textual labels that uniquely identify the controls or statements that the component implements.',  # noqa
            'Namespace': 'A namespace qualifying the property\'s name. This allows different organizations to associate distinct semantics with the same name. Used in conjunction with "class" as the ontology concept.',  # noqa
        }


class CsvRowMgr:
    """Csv row manager."""

    def __init__(self, row_names: List) -> None:
        """Initialize."""
        self.row_names = row_names
        self.map = {}

    def put(self, key: str, val: str) -> None:
        """Put."""
        if key not in self.row_names:
            raise RuntimeError(f'{key} not found')
        self.map[key] = val

    def get(self) -> List[str]:
        """Get."""
        row = []
        for name in self.row_names:
            if name in self.map.keys():
                row.append(self.map[name])
            else:
                row.append('')
        return row


class CisControlsHelper:
    """Cis controls helper."""

    def __init__(self, cis_controls: str) -> None:
        """Initialize."""
        self.cis_controls = cis_controls

    def ctl_generator(self) -> Iterator[Dict]:
        """Generate ctls until finished."""
        parts = self.cis_controls.split(';TITLE:')
        # restore TITLE:
        for n in range(len(parts)):
            parts[n] = f'TITLE:{parts[n]}'
        # process triples TITLE, CONTROL, DESCRIPTION
        for part in parts:
            if part:
                s1 = part.split('DESCRIPTION:')
                description = s1[1].strip()
                s2 = s1[0].split('CONTROL:')
                control = s2[1].strip()
                control_version = control.split()[0]
                control_id = control.split()[1]
                s3 = s2[0].split('TITLE:')
                title = s3[1].strip()
                ctl = {
                    'description': description,
                    'control-version': control_version,
                    'control-id': control_id,
                    'title': title
                }
                yield ctl

    def get_ctl_list(self, ctl_pfx: str, profile_version: List[str]) -> List[str]:
        """Get control list."""
        ctl_list = []
        if self.cis_controls:
            for ctl in self.ctl_generator():
                if ctl['control-version'] not in profile_version:
                    continue
                ctl_id = ctl['control-id']
                try:
                    float(ctl_id)
                except Exception:
                    text = f'missing or invalid control-id: "{ctl_id}"'
                    raise RuntimeError(text)
                ctl_list.append(f'{ctl_pfx}{ctl_id}')
        return ctl_list


class XlsxToCsvHelper:
    """Xlsx to csv helper."""

    def __init__(self, config_object: SectionProxy, tmpdir: str) -> None:
        """Initialize."""
        self.config_object = config_object
        benchmark_file = self.config_object['benchmark-file']
        self.ipath = pathlib.Path(benchmark_file)
        self.xpath = pathlib.Path(tmpdir) / self.ipath.name
        path = pathlib.Path(tmpdir) / self.xpath.name
        self.opath = path.with_suffix('.csv')
        self.wb = load_workbook(self.xpath)
        # worksheet
        self.ws = self.wb[SheetHelper.get_sheetname()]
        self._create_maps()
        # excluded columns
        default_columns_exclude = [f'"{head_recommendation_no}"', '"Profile"', '"Description"']
        columns_exclude = self.config_object.get('columns-exclude')
        if columns_exclude:
            columns_exclude = columns_exclude.strip().split(',')
        else:
            columns_exclude = default_columns_exclude
        self._columns_exclude = []
        for col in columns_exclude:
            name = PropertyHelper.normalize_name(col.replace('"', ''))
            self._columns_exclude.append(name)
        # benchmark control prefix
        self._benchmark_control_prefix = self.config_object.get(
            'benchmark-control-prefix', default_benchmark_control_prefix
        )
        if not self._benchmark_control_prefix.endswith('-'):
            self._benchmark_control_prefix = f'{self._benchmark_control_prefix}-'
        # benchmark rule prefix
        self._benchmark_rule_prefix = self.config_object.get('benchmark-rule-prefix', default_benchmark_rule_prefix)
        if not self._benchmark_rule_prefix.endswith('-'):
            self._benchmark_rule_prefix = f'{self._benchmark_rule_prefix}-'

    def _create_maps(self) -> Dict:
        """Create maps."""
        self._map_col_key_to_number = {}
        self._map_name_to_col_key = {}
        row = 1
        cols = self.ws.max_column + 1
        for col in range(row, cols):
            cell = self.ws.cell(row, col)
            if cell.value:
                key = self._name_to_key(cell.value)
                self._map_col_key_to_number[key] = col
                self._map_name_to_col_key[cell.value] = key

    def _name_to_key(self, name: str) -> str:
        """Name to key."""
        rval = name.lower()
        return rval

    def _sanitize(self, value: str) -> str:
        """Sanitize value."""
        rval = value
        if value:
            rval = value.replace('\n', ' ')
        return rval

    def _get_map(self) -> Dict:
        """Get map."""
        return self._map_name_to_col_key

    def get_all_columns(self) -> List:
        """Get all columns."""
        return self._get_map().keys()

    def row_generator(self) -> Iterator[int]:
        """Generate rows until max reached."""
        row = 2
        while row <= self.ws.max_row:
            yield row
            row += 1

    def get(self, row: int, name: str) -> str:
        """Get cell value for given row and column name."""
        key = self._name_to_key(name)
        col = self._map_col_key_to_number[key]
        cell = self.ws.cell(row, col)
        return self._sanitize(cell.value)

    def is_same_rule(self, row_a: int, row_b: str) -> str:
        """Is same rule."""
        rule_a = self.get(row_a, head_recommendation_no)
        rule_b = self.get(row_b, head_recommendation_no)
        rval = rule_a == rule_b
        logger.debug(f'{rval} {rule_a} {rule_b}')
        return rval

    def is_excluded_column(self, column: str) -> bool:
        """Is excluded column."""
        for item in self._columns_exclude:
            if item.lower() == column.lower():
                return True
        return False

    def merge_row(self, prev_row: int, curr_row: int) -> bool:
        """Merge row."""
        rval = False
        if prev_row and self.is_same_rule(prev_row, curr_row):
            prof2 = self.get(prev_row, 'Profile')
            prof1 = self.get(curr_row, 'Profile')
            prof = f'"{prof1}; {prof2}"'
            self.csv_helper.delete_last_row()
            # col Profile
            self.csv_row_mgr.put('Profile', prof)
            # add body row
            row_body = self.csv_row_mgr.get()
            self.csv_helper.add_row(row_body)
            rval = True
        return rval

    def _is_column(self, c1: str, c2: str) -> bool:
        """Is column."""
        if c1.lower() == c2.lower():
            rval = True
        else:
            rval = False
        return rval

    def heading_row_1(self) -> None:
        """Heading row 1."""
        self.row_names = []
        for col_name in CsvHelper.columns().keys():
            name = PropertyHelper.normalize_name(col_name)
            self.row_names.append(name)

    def heading_row_2(self) -> None:
        """Heading row 2."""
        self.row_descs = []
        for col_desc in CsvHelper.columns().values():
            desc = col_desc
            self.row_descs.append(desc)
        # additional user columns
        for col in self.get_all_columns():
            name = PropertyHelper.normalize_name(col)
            if self.is_excluded_column(col):
                continue
            self.row_names.append(name)
            self.row_descs.append(name)
        # additional non-rule columns
        for col in self.non_rule_helper.get_all_columns():
            name = PropertyHelper.normalize_name(col)
            self.row_names.append(name)
            self.row_descs.append(name)

    def _get_ctl_list(self, prev_row: int, curr_row: int) -> List[str]:
        """Get_ctl_list."""
        ctl_list = []
        # if merged row, list is empty
        if not self.merge_row(prev_row, curr_row):
            # if non-rule row, list is empty
            rec_no = self.get(curr_row, head_recommendation_no)
            if rec_no is not None:
                # get list
                cis_controls = self.get(curr_row, 'CIS Controls')
                cis_control_helper = CisControlsHelper(cis_controls)
                ctl_list = cis_control_helper.get_ctl_list(
                    self._benchmark_control_prefix, self.config_object['profile-version']
                )
        return ctl_list

    def run(self) -> None:
        """Run."""
        self.csv_helper = CsvHelper(self.opath)
        self.non_rule_helper = NonRuleHelper(self.config_object, self)
        # heading row 1 - names
        self.heading_row_1()
        # heading row 2 - descriptions
        self.heading_row_2()
        # add heading rows
        self.csv_helper.add_row(self.row_names)
        self.csv_helper.add_row(self.row_descs)
        # body
        user_columns = []
        for col in self.get_all_columns():
            if self.is_excluded_column(col):
                continue
            user_columns.append(col)
        # process each data row of CIS Benchmark file
        prev_row = None
        for curr_row in self.row_generator():
            ctl_list = self._get_ctl_list(prev_row, curr_row)
            if not ctl_list:
                continue
            # create new row
            self.csv_row_mgr = CsvRowMgr(self.row_names)
            # col Component_Title
            self.csv_row_mgr.put('Component_Title', self.config_object['component-name'])
            # col Component_Description
            self.csv_row_mgr.put('Component_Description', self.config_object['component-description'])
            # col Component_Type
            self.csv_row_mgr.put('Component_Type', self.config_object['component-type'])
            # col Rule_ID
            rec_no = self.get(curr_row, head_recommendation_no)
            rule_id = f'{self._benchmark_rule_prefix}{rec_no}'
            self.csv_row_mgr.put('Rule_Id', rule_id)
            # col Rule_Description
            rule_desc = self.get(curr_row, 'Description')
            self.csv_row_mgr.put('Rule_Description', rule_desc)
            # col Profile_Source
            self.csv_row_mgr.put('Profile_Source', self.config_object['profile-source'])
            # col Profile_Description
            self.csv_row_mgr.put('Profile_Description', self.config_object['profile-description'])
            # col Control_Id_List
            ctl_list = ' '.join(ctl_list)
            self.csv_row_mgr.put('Control_Id_List', ctl_list)
            # col Namespace
            self.csv_row_mgr.put('Namespace', self.config_object['namespace'])
            # col Profile
            prof = self.get(curr_row, 'Profile')
            self.csv_row_mgr.put('Profile', prof)
            # col user
            for col in user_columns:
                val = self.get(curr_row, col)
                if val:
                    if self._is_column(col, 'cis controls'):
                        val = f'"{val}"'
                    name = PropertyHelper.normalize_name(col)
                    self.csv_row_mgr.put(name, val)
            # col group levels
            self.non_rule_helper.add_group_levels(rec_no, self.csv_row_mgr)
            # add body row
            row_body = self.csv_row_mgr.get()
            self.csv_helper.add_row(row_body)
            # previous row
            prev_row = curr_row
        # write body
        self.csv_helper.save()


class NonRuleHelper:
    """Non-rule Helper."""

    def __init__(self, config: SectionProxy, xlsx_helper: XlsxToCsvHelper) -> None:
        """Initialize."""
        self.col_prefix = 'Group'
        self.col_level = 'Level'
        self.config = config
        self.xlsx_helper = xlsx_helper
        default_columns_carry_forward = [head_section_no, 'Title', 'Description']
        self.columns_carry_forward = self.config.get('columns-carry-forward', default_columns_carry_forward)
        self.col_list = []
        self.sec_map = {}
        if self.columns_carry_forward:
            self._init_col_list()
            self._init_sec_map()

    def _init_col_list(self) -> None:
        """Init col list."""
        biggest = -1
        for row in self.xlsx_helper.row_generator():
            rec_no = self.xlsx_helper.get(row, head_recommendation_no)
            if rec_no:
                continue
            sec_no = self.xlsx_helper.get(row, head_section_no)
            count = sec_no.count('.')
            if count > biggest:
                biggest = count
        for i in range(biggest + 1):
            for col_name in self.columns_carry_forward:
                name = f'{self.col_prefix}_{col_name}_{self.col_level}_{i}'
                name = PropertyHelper.normalize_name(name)
                self.col_list.append(name)

    def _init_sec_map(self) -> None:
        """Init sec map."""
        for row in self.xlsx_helper.row_generator():
            rec_no = self.xlsx_helper.get(row, head_recommendation_no)
            if rec_no:
                continue
            sec_no = self.xlsx_helper.get(row, head_section_no)
            _map = {}
            for name in self.columns_carry_forward:
                _map[name] = self.xlsx_helper.get(row, name)
            self.sec_map[sec_no] = _map

    def get_all_columns(self) -> List:
        """Get all columns."""
        return self.col_list

    def add_group_levels(self, rec_no: str, csv_row_mgr: CsvRowMgr) -> None:
        """Add group levels."""
        parts = rec_no.split('.')[:-1]
        for i, part in enumerate(parts):
            if not i:
                sec_no = f'{part}'
            else:
                sec_no = f'{sec_no}.{part}'
            if sec_no in self.sec_map.keys():
                _map = self.sec_map[sec_no]
                for key in _map.keys():
                    val = _map[key]
                    name = f'{self.col_prefix}_{key}_{self.col_level}_{i}'
                    name = PropertyHelper.normalize_name(name)
                    csv_row_mgr.put(name, val)


class CsvToJsonHelper:
    """Csv to json helper."""

    def __init__(self, config_object: Optional[configparser.SectionProxy], tmpdir: str) -> None:
        """Initialize trestle task."""
        self.config_object = config_object
        benchmark_file = self.config_object['benchmark-file']
        self.ipath = pathlib.Path(benchmark_file)
        self.xpath = pathlib.Path(tmpdir) / self.ipath.name
        path = pathlib.Path(tmpdir) / self.xpath.name
        self.opath = path.with_suffix('.csv')
        self.config_object['csv-file'] = str(self.opath)
        self.csv_to_oscal_cd = CsvToOscalComponentDefinition(self.config_object)
        self.config_object['title'] = self.config_object['benchmark-title']
        self.config_object['version'] = self.config_object['benchmark-version']

    def run(self) -> None:
        """Run."""
        return self.csv_to_oscal_cd.execute()
