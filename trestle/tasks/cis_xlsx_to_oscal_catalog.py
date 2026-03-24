# -*- mode:python; coding:utf-8 -*-
# Copyright (c) 2023 IBM Corp. All rights reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     https://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
"""OSCAL transformation tasks."""

# mypy: ignore-errors  # noqa E800
import configparser
import datetime
import logging
import pathlib
import traceback
import uuid
from collections import OrderedDict
from typing import Iterator, List, Optional

from openpyxl import load_workbook

from trestle.oscal import OSCAL_VERSION
from trestle.oscal.catalog import Catalog
from trestle.oscal.catalog import Control
from trestle.oscal.catalog import Group1 as CatalogGroup1
from trestle.oscal.catalog import Group2 as CatalogGroup2
from trestle.oscal.common import BackMatter
from trestle.oscal.common import Link
from trestle.oscal.common import Metadata
from trestle.oscal.common import Part
from trestle.oscal.common import Property
from trestle.oscal.common import Resource
from trestle.tasks.base_task import TaskBase
from trestle.tasks.base_task import TaskOutcome

logger = logging.getLogger(__name__)

timestamp = datetime.datetime.now(datetime.timezone.utc).replace(microsecond=0).isoformat()


class XlsxHelper:
    """Xlsx Helper common functions and assistance navigating spread sheet."""

    def __init__(self, file: str) -> None:
        """Initialize."""
        self._spread_sheet = file
        self._wb = load_workbook(self._spread_sheet)
        sheet_candidates = ['Combined Profiles', 'Combined']
        self._sheet_name = None
        for sheet_candidate in sheet_candidates:
            if sheet_candidate in self._wb.sheetnames:
                self._sheet_name = sheet_candidate
                break
        if not self._sheet_name:
            raise RuntimeError(f'{file} missing one of {sheet_candidates} sheet')
        self._work_sheet = self._wb[self._sheet_name]
        self._mapper()
        self._key_to_col_map = {'statement': 'description'}

    def is_ocp(self) -> bool:
        """Check if sheet is for OCP."""
        return self._sheet_name == 'Combined Profiles'

    def _normalize(self, name: str) -> str:
        """Normalize."""
        return name.lower()

    def _translate(self, name: str) -> str:
        """Translate name key to column name."""
        return self._key_to_col_map.get(name, name)

    def _mapper(self) -> None:
        """Map columns heading names to column numbers."""
        self._col_name_to_number = {}
        cols = self._work_sheet.max_column + 1
        row = 1
        for col in range(row, cols):
            cell = self._work_sheet.cell(row, col)
            if cell.value:
                name = self._normalize(cell.value)
                self._col_name_to_number[name] = col

    def row_generator(self) -> Iterator[int]:
        """Generate rows until max reached."""
        row = 2
        while row <= self._work_sheet.max_row:
            yield row
            row += 1

    def get(self, row: int, name: str) -> str:
        """Get cell value for given row and column name."""
        nname = self._normalize(name)
        cname = self._translate(nname)
        col = self._col_name_to_number[cname]
        cell = self._work_sheet.cell(row, col)
        return cell.value


class CatalogHelper:
    """OSCAL Catalog Helper."""

    def __init__(self, title: str, version: str) -> None:
        """Initialize."""
        # metadata
        self._metadata = Metadata(title=title, last_modified=timestamp, oscal_version=OSCAL_VERSION, version=version)
        self._root_group = OrderedDict()
        self._root_resources = OrderedDict()
        self._all_groups = OrderedDict()
        self._all_controls = OrderedDict()

    def add_group(self, section: str, title: str, props: List[Property], parts: List[Part]) -> None:
        """Add group - always creates Group1 initially (Pass 1)."""
        numdots = section.count('.')
        if numdots == 0:
            group = CatalogGroup1(title=f'{title}', id=f'CIS-{section}')
            if props:
                group.props = props
            if parts:
                group.parts = parts
            self._root_group[section] = group
            self._all_groups[section] = group
        else:
            key = '.'.join(section.split('.')[:-1])
            parent = self._all_groups[key]
            # Parent must be Group1 at this point (Pass 1 - only groups added)
            if not hasattr(parent, 'groups'):
                raise RuntimeError(
                    f'Parent {key} is not Group1 when adding group {section} - this should not happen in Pass 1'
                )
            if parent.groups is None:
                parent.groups = []
            group = CatalogGroup1(title=f'{title}', id=f'CIS-{section}')
            if props:
                group.props = props
            if parts:
                group.parts = parts
            parent.groups.append(group)
            self._all_groups[section] = group

    def prepare_groups_for_controls(self, sections_with_controls: set) -> None:
        """Prepare groups that will have controls (Pass 1.5).

        For groups that will have controls AND have subgroups, create a controls subgroup.
        """
        for section in sections_with_controls:
            if section in self._all_groups:
                group = self._all_groups[section]
                # Check if this is a Group1 with subgroups
                if hasattr(group, 'groups') and group.groups:
                    # Create a "controls" subgroup (Group2) for the controls
                    controls_group_id = f'{group.id}-controls'
                    controls_group = CatalogGroup2(id=controls_group_id, title=f'{group.title} Controls')
                    controls_group.controls = []
                    group.groups.append(controls_group)
                    # Store reference for quick access in Pass 2
                    self._all_groups[f'{section}_controls'] = controls_group

    def _add_prop(self, control: Control, prop: Property) -> None:
        """Add property to control."""
        control_props = control.props
        control.props = []
        last = 0
        for i, control_prop in enumerate(control_props):
            if control_prop.name == prop.name:
                last = i
        for i, control_prop in enumerate(control_props):
            control.props.append(control_prop)
            if i == last:
                control.props.append(prop)

    def add_control(
        self, section: str, recommendation: str, title: str, props: List[Property], parts: List[Part], links: List[Link]
    ) -> None:
        """Add control (Pass 2) - uses prepared group structure."""
        # Check if there's a controls subgroup for this section
        controls_key = f'{section}_controls'
        if controls_key in self._all_groups:
            # Use the controls subgroup created in Pass 1.5
            group = self._all_groups[controls_key]
        else:
            # No controls subgroup - this group has no other subgroups
            group = self._all_groups[section]

            # If it's still Group1, convert to Group2
            if not hasattr(group, 'controls'):
                group2 = CatalogGroup2(id=group.id, title=group.title)
                if hasattr(group, 'props') and group.props:
                    group2.props = group.props
                if hasattr(group, 'parts') and group.parts:
                    group2.parts = group.parts
                if hasattr(group, 'links') and group.links:
                    group2.links = group.links
                # Replace in parent or root
                if section.count('.') == 0:
                    self._root_group[section] = group2
                else:
                    key = '.'.join(section.split('.')[:-1])
                    parent = self._all_groups[key]
                    if hasattr(parent, 'groups') and parent.groups:
                        for i, g in enumerate(parent.groups):
                            if g.id == group.id:
                                parent.groups[i] = group2
                                break
                self._all_groups[section] = group2
                group = group2

        if group.controls is None:
            group.controls = []
        id_ = f'CIS-{recommendation}'
        if id_ in self._all_controls:
            control = self._all_controls[id_]
            for prop in props:
                if prop.name == 'profile':
                    self._add_prop(control, prop)
        else:
            title = f'{title}'
            control = Control(id=id_, title=title)
            self._all_controls[id_] = control
            if props:
                control.props = props
            if parts:
                control.parts = parts
            if links:
                control.links = links
            group.controls.append(control)

    def add_link(self, recommendation: str, reference: str, links: List[Link]) -> None:
        """Add link."""
        id_ = f'CIS-{recommendation}'
        if id_ not in self._root_resources:
            res_id = str(uuid.uuid4())
            link = Link(href=f'#{res_id}', rel='reference')
            links.append(link)
            resource = Resource(uuid=res_id, description=reference)
            self._root_resources[id_] = resource

    def get_catalog(self) -> Catalog:
        """Get catalog."""
        back_matter = BackMatter(resources=list(self._root_resources.values()))
        catalog = Catalog(
            uuid=str(uuid.uuid4()),
            metadata=self._metadata,
            groups=list(self._root_group.values()),
            back_matter=back_matter,
        )
        return catalog


class CisXlsxToOscalCatalog(TaskBase):
    """
    Task to transform CIS .xlsx to OSCAL catalog.

    Attributes:
        name: Name of the task.
    """

    name = 'cis-xlsx-to-oscal-catalog'
    ns = 'https://oscal-compass.github.io/compliance-trestle/schemas/oscal/catalog/cis'

    def __init__(self, config_object: Optional[configparser.SectionProxy]) -> None:
        """
        Initialize trestle task.

        Args:
            config_object: Config section associated with the task.
        """
        super().__init__(config_object)
        logger.warning('This task is deprecated. Use task cis-xlsx-to-oscal-cd.')

    def print_info(self) -> None:
        """Print the help string."""
        logger.info(f'Help information for {self.name} task.')
        logger.info('')
        logger.info('Purpose: Create catalog from standard (e.g. CIS benchmark).')
        logger.info('')
        logger.info('Configuration flags sit under [task.cis-xlsx-to-oscal-catalog]:')
        text1 = '  input-file             = '
        text2 = '(required) path to read the compliance-as-code .xlsx spread sheet file.'
        logger.info(text1 + text2)
        text1 = '  output-dir             = '
        text2 = '(required) location to write the generated catalog.json file.'
        logger.info(text1 + text2)
        text1 = '  title                  = '
        text2 = '(required) title of the CIS catalog.'
        logger.info(text1 + text2)
        text1 = '  version                = '
        text2 = '(required) version of the CIS catalog.'
        logger.info(text1 + text2)
        text1 = '  output-overwrite       = '
        text2 = '(optional) true [default] or false; replace existing output when true.'
        logger.info(text1 + text2)

    def simulate(self) -> TaskOutcome:
        """Provide a simulated outcome."""
        return TaskOutcome('simulated-success')

    def execute(self) -> TaskOutcome:
        """Provide an actual outcome."""
        try:
            return self._execute()
        except Exception:
            logger.info(traceback.format_exc())
            return TaskOutcome('failure')

    def _get_normalized_name(self, key: str) -> str:
        """Get normalized name."""
        name = key
        name = name.replace(' ', '_')
        name = name.replace('&', 'a')
        name = name.replace('(', '')
        name = name.replace(')', '')
        return name

    def _create_property(self, name: str, value: str) -> Property:
        return Property(name=name, value=value, ns=CisXlsxToOscalCatalog.ns)

    def _add_property(self, xlsx_helper: XlsxHelper, props: List[Property], row: int, key: str) -> None:
        """Add property."""
        try:
            name = self._get_normalized_name(key)
            value = xlsx_helper.get(row, key)
            if value:
                props.append(self._create_property(name, value))
        except KeyError as e:
            text = f'key {e} not found.'
            logger.debug(text)

    def _add_property_boolean(self, xlsx_helper: XlsxHelper, props: List[Property], row: int, key: str) -> None:
        """Add property."""
        try:
            name = self._get_normalized_name(key)
            value = xlsx_helper.get(row, key)
            if value:
                props.append(self._create_property(name, 'True'))
            else:
                props.append(self._create_property(name, 'False'))
        except KeyError as e:
            text = f'key {e} not found.'
            logger.debug(text)

    def _add_part(self, xlsx_helper: XlsxHelper, parts: List[Part], id_: str, row: int, key: str) -> None:
        """Add part."""
        value = xlsx_helper.get(row, key)
        if value:
            name = self._get_normalized_name(key)
            parts.append(Part(id=id_, name=name, prose=value))

    def _add_links(
        self, xlsx_helper: XlsxHelper, catalog_helper: CatalogHelper, links: List[Link], row: int, key: str
    ) -> None:
        """Add links."""
        value = xlsx_helper.get(row, key)
        if value:
            recommendation = xlsx_helper.get(row, 'recommendation #')
            catalog_helper.add_link(recommendation, value, links)

    def _execute(self) -> TaskOutcome:
        """Wrap the execute for exception handling."""
        if not self._config:
            logger.warning('config missing')
            return TaskOutcome('failure')
        try:
            ifile = self._config['input-file']
            odir = self._config['output-dir']
            title = self._config['title']
            version = self._config['version']
        except KeyError as e:
            logger.info(f'key {e.args[0]} missing')
            return TaskOutcome('failure')
        # verbosity
        _quiet = self._config.get('quiet', False)
        _verbose = not _quiet
        # output
        overwrite = self._config.getboolean('output-overwrite', True)
        opth = pathlib.Path(odir)
        # insure output dir exists
        opth.mkdir(exist_ok=True, parents=True)
        # calculate output file name & check writability
        oname = 'catalog.json'
        ofile = opth / oname
        if not overwrite and pathlib.Path(ofile).exists():
            logger.warning(f'output: {ofile} already exists')
            return TaskOutcome('failure')
        xlsx_helper = XlsxHelper(ifile)
        catalog_helper = CatalogHelper(title, version)
        if xlsx_helper.is_ocp():
            self._process_ocp(xlsx_helper, catalog_helper)
        else:
            self._process_rhel(xlsx_helper, catalog_helper)
        catalog = catalog_helper.get_catalog()
        # write OSCAL ComponentDefinition to file
        if _verbose:
            logger.info(f'output: {ofile}')
        catalog.oscal_write(pathlib.Path(ofile))
        return TaskOutcome('success')

    def _process_ocp(self, xlsx_helper: XlsxHelper, catalog_helper: CatalogHelper) -> None:
        """Process OCP with three-pass approach for OSCAL 1.2.0."""
        # Pre-scan: Determine which sections will have controls
        sections_with_controls = set()
        for row in xlsx_helper.row_generator():
            recommendation = xlsx_helper.get(row, 'recommendation #')
            if recommendation is not None:
                section = xlsx_helper.get(row, 'section #')
                sections_with_controls.add(section)

        # Pass 1: Build complete group hierarchy (all Group1)
        for row in xlsx_helper.row_generator():
            section = xlsx_helper.get(row, 'section #')
            recommendation = xlsx_helper.get(row, 'recommendation #')
            title = xlsx_helper.get(row, 'title')

            if recommendation is None:
                # This is a group/section
                props = []
                parts = []
                frag = section
                # props
                self._add_property(xlsx_helper, props, row, 'profile')
                self._add_property(xlsx_helper, props, row, 'status')
                self._add_property(xlsx_helper, props, row, 'assessment status')
                # parts
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_smt', row, 'statement')
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_rat', row, 'rationale statement')
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_imp', row, 'impact statement')
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_rem', row, 'remediation procedure')
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_aud', row, 'audit procedure')
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_inf', row, 'additional information')
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_ctl', row, 'CIS Controls')
                catalog_helper.add_group(section, title, props, parts)

        # Pass 1.5: Prepare groups that will have controls
        catalog_helper.prepare_groups_for_controls(sections_with_controls)

        # Pass 2: Add controls
        for row in xlsx_helper.row_generator():
            section = xlsx_helper.get(row, 'section #')
            recommendation = xlsx_helper.get(row, 'recommendation #')
            title = xlsx_helper.get(row, 'title')

            if recommendation is not None:
                # This is a control
                props = []
                parts = []
                links = []
                frag = recommendation
                # props
                self._add_property(xlsx_helper, props, row, 'profile')
                self._add_property(xlsx_helper, props, row, 'status')
                self._add_property(xlsx_helper, props, row, 'assessment status')
                # parts
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_smt', row, 'statement')
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_rat', row, 'rationale statement')
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_imp', row, 'impact statement')
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_rem', row, 'remediation procedure')
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_aud', row, 'audit procedure')
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_inf', row, 'additional information')
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_ctl', row, 'CIS Controls')
                self._add_property_boolean(xlsx_helper, props, row, 'v7 IG1')
                self._add_property_boolean(xlsx_helper, props, row, 'v7 IG2')
                self._add_property_boolean(xlsx_helper, props, row, 'v7 IG3')
                self._add_property_boolean(xlsx_helper, props, row, 'v8 IG1')
                self._add_property_boolean(xlsx_helper, props, row, 'v8 IG2')
                self._add_property_boolean(xlsx_helper, props, row, 'v8 IG3')
                self._add_property(xlsx_helper, props, row, 'MITRE ATT&CK Mappings')
                self._add_links(xlsx_helper, catalog_helper, links, row, 'references')
                catalog_helper.add_control(section, recommendation, title, props, parts, links)

    def _process_rhel(self, xlsx_helper: XlsxHelper, catalog_helper: CatalogHelper) -> None:
        """Process RHEL with three-pass approach for OSCAL 1.2.0."""
        # Pre-scan: Determine which sections will have controls
        sections_with_controls = set()
        for row in xlsx_helper.row_generator():
            recommendation = xlsx_helper.get(row, 'recommendation #')
            if recommendation is not None:
                section = xlsx_helper.get(row, 'section #')
                sections_with_controls.add(section)

        # Pass 1: Build complete group hierarchy (all Group1)
        for row in xlsx_helper.row_generator():
            section = xlsx_helper.get(row, 'section #')
            recommendation = xlsx_helper.get(row, 'recommendation #')
            title = xlsx_helper.get(row, 'title')

            if recommendation is None:
                # This is a group/section
                props = []
                parts = []
                frag = section
                # props
                self._add_property(xlsx_helper, props, row, 'profile')
                self._add_property(xlsx_helper, props, row, 'assessment status')
                # parts
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_smt', row, 'statement')
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_rat', row, 'rational statement')
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_imp', row, 'impact statement')
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_rem', row, 'remediation procedure')
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_aud', row, 'audit procedure')
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_inf', row, 'additional information')
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_ctl', row, 'CIS Controls')
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_1v8', row, 'CIS Safeguards 1 (v8)')
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_2v8', row, 'CIS Safeguards 2 (v8)')
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_3v8', row, 'CIS Safeguards 3 (v8)')
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_1v7', row, 'CIS Safeguards 1 (v7)')
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_2v7', row, 'CIS Safeguards 2 (v7)')
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_3v7', row, 'CIS Safeguards 3 (v7)')
                catalog_helper.add_group(section, title, props, parts)

        # Pass 1.5: Prepare groups that will have controls
        catalog_helper.prepare_groups_for_controls(sections_with_controls)

        # Pass 2: Add controls
        for row in xlsx_helper.row_generator():
            section = xlsx_helper.get(row, 'section #')
            recommendation = xlsx_helper.get(row, 'recommendation #')
            title = xlsx_helper.get(row, 'title')

            if recommendation is not None:
                # This is a control
                props = []
                parts = []
                links = []
                frag = recommendation
                # props
                self._add_property(xlsx_helper, props, row, 'profile')
                self._add_property(xlsx_helper, props, row, 'assessment status')
                # parts
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_smt', row, 'statement')
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_rat', row, 'rational statement')
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_imp', row, 'impact statement')
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_rem', row, 'remediation procedure')
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_aud', row, 'audit procedure')
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_inf', row, 'additional information')
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_ctl', row, 'CIS Controls')
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_1v8', row, 'CIS Safeguards 1 (v8)')
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_2v8', row, 'CIS Safeguards 2 (v8)')
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_3v8', row, 'CIS Safeguards 3 (v8)')
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_1v7', row, 'CIS Safeguards 1 (v7)')
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_2v7', row, 'CIS Safeguards 2 (v7)')
                self._add_part(xlsx_helper, parts, f'CIS-{frag}_3v7', row, 'CIS Safeguards 3 (v7)')
                self._add_property_boolean(xlsx_helper, props, row, 'v7 IG1')
                self._add_property_boolean(xlsx_helper, props, row, 'v7 IG2')
                self._add_property_boolean(xlsx_helper, props, row, 'v7 IG3')
                self._add_property_boolean(xlsx_helper, props, row, 'v8 IG1')
                self._add_property_boolean(xlsx_helper, props, row, 'v8 IG2')
                self._add_property_boolean(xlsx_helper, props, row, 'v8 IG3')
                self._add_links(xlsx_helper, catalog_helper, links, row, 'references')
                catalog_helper.add_control(section, recommendation, title, props, parts, links)
