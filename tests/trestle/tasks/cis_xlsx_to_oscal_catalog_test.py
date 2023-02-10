# -*- mode:python; coding:utf-8 -*-
# Copyright (c) 2023 IBM Corp. All rights reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     https://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
"""cis-xlsx-to-oscal-catalog task tests."""

import configparser
import os
import pathlib
from typing import Dict
from unittest import mock

from openpyxl import load_workbook

import trestle.tasks.cis_xlsx_to_oscal_catalog as cis_xlsx_to_oscal_catalog
from trestle.oscal.catalog import Catalog
from trestle.tasks.base_task import TaskOutcome

ocp_config = 'tests/data/tasks/cis-xlsx-to-oscal-catalog/test-cis-xlsx-to-oscal-catalog.ocp.config'
rhel_config = 'tests/data/tasks/cis-xlsx-to-oscal-catalog/test-cis-xlsx-to-oscal-catalog.rhel.config'


def _get_section(tmp_path: pathlib.Path, file_: str) -> Dict:
    """Get section."""
    config = configparser.ConfigParser()
    config_path = pathlib.Path(file_)
    config.read(config_path)
    section = config['task.cis-xlsx-to-oscal-catalog']
    section['output-dir'] = str(tmp_path)
    return section


def test_cis_xlsx_to_oscal_catalog_print_info(tmp_path: pathlib.Path):
    """Test print_info call."""
    section = _get_section(tmp_path, ocp_config)
    tgt = cis_xlsx_to_oscal_catalog.CisXlsxToOscalCatalog(section)
    retval = tgt.print_info()
    assert retval is None


def test_cis_xlsx_to_oscal_catalog_simulate(tmp_path: pathlib.Path):
    """Test simulate call."""
    section = _get_section(tmp_path, ocp_config)
    tgt = cis_xlsx_to_oscal_catalog.CisXlsxToOscalCatalog(section)
    retval = tgt.simulate()
    assert retval == TaskOutcome.SIM_SUCCESS
    assert len(os.listdir(str(tmp_path))) == 0


def test_cis_xlsx_to_oscal_catalog_execute_ocp(tmp_path: pathlib.Path):
    """Test execute call - ocp."""
    section = _get_section(tmp_path, ocp_config)
    tgt = cis_xlsx_to_oscal_catalog.CisXlsxToOscalCatalog(section)
    retval = tgt.execute()
    assert retval == TaskOutcome.SUCCESS
    _validate_ocp(tmp_path)


def _validate_ocp(tmp_path: pathlib.Path):
    """Validate produced OSCAL for OCP catalog."""
    # read catalog
    file_path = tmp_path / 'catalog.json'
    catalog = Catalog.oscal_read(file_path)
    # spot check
    assert len(catalog.groups) == 1
    # group 0
    g0 = catalog.groups[0]
    assert g0.id == 'CIS-1'
    assert g0.title == 'Control Plane Components'
    assert len(g0.groups) == 2
    assert len(g0.props) == 1
    p0 = g0.props[0]
    assert p0.name == 'status'
    assert p0.value == 'draft'
    g00 = g0.groups[0]
    assert g00.id == 'CIS-1.1'
    assert g00.title == 'Master Node Configuration Files'
    assert len(g0.props) == 1
    assert len(g00.controls) == 7
    c5 = g00.controls[5]
    p50 = c5.props[0]
    assert p50.name == 'profile'
    assert p50.value == 'Level 1'
    p52 = c5.props[2]
    assert p52.name == 'assessment_status'
    assert p52.value == 'Manual'


def test_cis_xlsx_to_oscal_catalog_no_overwrite(tmp_path: pathlib.Path):
    """Test execute call."""
    section = _get_section(tmp_path, ocp_config)
    tgt = cis_xlsx_to_oscal_catalog.CisXlsxToOscalCatalog(section)
    retval = tgt.execute()
    assert retval == TaskOutcome.SUCCESS
    _validate_ocp(tmp_path)
    section['output-overwrite'] = 'False'
    tgt = cis_xlsx_to_oscal_catalog.CisXlsxToOscalCatalog(section)
    retval = tgt.execute()
    assert retval == TaskOutcome.FAILURE


def test_cis_xlsx_to_oscal_catalog_missing_config(tmp_path: pathlib.Path):
    """Test missing config."""
    section = None
    tgt = cis_xlsx_to_oscal_catalog.CisXlsxToOscalCatalog(section)
    retval = tgt.execute()
    assert retval == TaskOutcome.FAILURE


def test_cis_xlsx_to_oscal_catalog_missing_version(tmp_path: pathlib.Path):
    """Test missing version."""
    section = _get_section(tmp_path, ocp_config)
    section.pop('version')
    tgt = cis_xlsx_to_oscal_catalog.CisXlsxToOscalCatalog(section)
    retval = tgt.execute()
    assert retval == TaskOutcome.FAILURE


def test_cis_xlsx_to_oscal_catalog_missing_sheet(tmp_path: pathlib.Path):
    """Test missing sheet."""
    folder = 'tests/data/tasks/cis-xlsx-to-oscal-catalog'
    file_ = f'{folder}/CIS_RedHat_OpenShift_Container_Platform_Benchmark_v1.2.0-2.snippet.xlsx'
    wb_hacked = load_workbook(file_)
    sheet = wb_hacked['Combined Profiles']
    wb_hacked.remove(sheet)
    with mock.patch('trestle.tasks.cis_xlsx_to_oscal_catalog.load_workbook') as load_workbook_mock:
        load_workbook_mock.return_value = wb_hacked
        section = _get_section(tmp_path, ocp_config)
        tgt = cis_xlsx_to_oscal_catalog.CisXlsxToOscalCatalog(section)
        retval = tgt.execute()
        assert retval == TaskOutcome.FAILURE


def test_cis_xlsx_to_oscal_catalog_one_dot_added_part(tmp_path: pathlib.Path):
    """Test group part."""
    folder = 'tests/data/tasks/cis-xlsx-to-oscal-catalog'
    file_ = f'{folder}/CIS_RedHat_OpenShift_Container_Platform_Benchmark_v1.2.0-2.snippet.xlsx'
    wb_hacked = load_workbook(file_)
    sheet = wb_hacked['Combined Profiles']
    cell = sheet.cell(2, 7)
    assert cell.value.startswith('This section consists of security recommendations for')
    cell = sheet.cell(3, 7)
    cell.value = 'foobar'
    with mock.patch('trestle.tasks.cis_xlsx_to_oscal_catalog.load_workbook') as load_workbook_mock:
        load_workbook_mock.return_value = wb_hacked
        section = _get_section(tmp_path, ocp_config)
        tgt = cis_xlsx_to_oscal_catalog.CisXlsxToOscalCatalog(section)
        retval = tgt.execute()
        assert retval == TaskOutcome.SUCCESS
    # read catalog
    file_path = tmp_path / 'catalog.json'
    catalog = Catalog.oscal_read(file_path)
    # spot check
    g0 = catalog.groups[0]
    g00 = g0.groups[0]
    assert len(g00.parts) == 1
    p000 = g00.parts[0]
    assert p000.prose == 'foobar'


def test_cis_xlsx_to_oscal_catalog_unexpected_section(tmp_path: pathlib.Path):
    """Test group part."""
    folder = 'tests/data/tasks/cis-xlsx-to-oscal-catalog'
    file_ = f'{folder}/CIS_RedHat_OpenShift_Container_Platform_Benchmark_v1.2.0-2.snippet.xlsx'
    wb_hacked = load_workbook(file_)
    sheet = wb_hacked['Combined Profiles']
    cell = sheet.cell(3, 1)
    assert cell.value == '1.1'
    cell.value = '1.2.3.4.5.6.7.8.9.0'
    with mock.patch('trestle.tasks.cis_xlsx_to_oscal_catalog.load_workbook') as load_workbook_mock:
        load_workbook_mock.return_value = wb_hacked
        section = _get_section(tmp_path, ocp_config)
        tgt = cis_xlsx_to_oscal_catalog.CisXlsxToOscalCatalog(section)
        retval = tgt.execute()
        assert retval == TaskOutcome.FAILURE


def test_cis_xlsx_to_oscal_catalog_execute_rhel(tmp_path: pathlib.Path):
    """Test execute call - rhel."""
    section = _get_section(tmp_path, rhel_config)
    tgt = cis_xlsx_to_oscal_catalog.CisXlsxToOscalCatalog(section)
    retval = tgt.execute()
    assert retval == TaskOutcome.SUCCESS
    _validate_rhel(tmp_path)


def _validate_rhel(tmp_path: pathlib.Path):
    """Validate produced OSCAL for RHEL catalog."""
    # read catalog
    file_path = tmp_path / 'catalog.json'
    catalog = Catalog.oscal_read(file_path)
    # spot check
    assert len(catalog.groups) == 1
    # group 0
    g0 = catalog.groups[0]
    assert g0.id == 'CIS-1'
    assert g0.title == 'Initial Setup'
    assert len(g0.groups) == 1
    assert len(g0.parts) == 1
    assert len(g0.controls) == 2
    p00 = g0.parts[0]
    assert p00.id == 'CIS-1_smt'
    assert p00.name == 'statement'
    assert p00.prose.startswith('Items in this section are advised for all systems,')
    g00 = g0.groups[0]
    assert g00.id == 'CIS-1.1'
    assert g00.title == 'Filesystem Configuration'
    c00 = g0.controls[0]
    assert c00.id == 'CIS-1.9'
    assert c00.title == 'Ensure updates, patches, and additional security software are installed'
    assert len(c00.props) == 9
    prop0 = c00.props[0]
    assert prop0.name == 'profile'
    assert prop0.value == 'Level 1 - Server'
    assert len(c00.parts) == 9
    part0 = c00.parts[0]
    assert part0.id == 'CIS-1.9_smt'
    assert part0.name == 'statement'
    t1 = 'Periodically patches are released for included software either due'
    t2 = ' to security flaws or to include additional functionality.'
    assert part0.prose == t1 + t2
    md = catalog.metadata
    assert md.title == 'CIS Red Hat Enterprise Linux 8 Benchmark'
    rs = catalog.back_matter.resources
    assert len(rs) == 1
    assert rs[
        0].description == 'CRYPTO-POLICIES(7):https://access.redhat.com/articles/3642912#what-polices-are-provided-1'
